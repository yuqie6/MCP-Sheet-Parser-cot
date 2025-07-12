import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

def create_failing_style_sample(file_path: str):
    """
    Creates an Excel file with complex styles to test the XlsxParser.
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None, "Sheet should not be None"
    sheet.title = "Complex Styles"

    # --- Define Styles ---
    style_1 = {
        "font": Font(name='Arial', size=12, bold=True, italic=True, color="FF0000"),
        "fill": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
        "border": Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='double', color='00FF00'),
            top=Side(style='dashed', color='0000FF'),
            bottom=Side(style='thick', color='FF00FF')
        ),
        "alignment": Alignment(horizontal='center', vertical='center', wrap_text=True)
    }

    style_2 = {
        "font": Font(name='Courier New', size=10, underline='single', color='0000FF'),
        "fill": PatternFill(fill_type="lightGray"),
        "alignment": Alignment(horizontal='right', vertical='bottom')
    }
    
    style_3 = {
        "font": Font(name='Times New Roman', size=16, color="E06666"),
        "border": Border(left=Side(style='medium'), right=Side(style='medium')),
        "alignment": Alignment(horizontal='left', vertical='top')
    }
    
    # --- Apply Styles to Cells ---
    sheet["A1"].value = "Style 1"
    sheet["A1"].font = style_1["font"]
    sheet["A1"].fill = style_1["fill"]
    sheet["A1"].border = style_1["border"]
    sheet["A1"].alignment = style_1["alignment"]

    sheet["B2"].value = "Style 2"
    sheet["B2"].font = style_2["font"]
    sheet["B2"].fill = style_2["fill"]
    sheet["B2"].alignment = style_2["alignment"]
    
    sheet["C3"].value = "Style 3"
    sheet["C3"].font = style_3["font"]
    sheet["C3"].border = style_3["border"]
    sheet["C3"].alignment = style_3["alignment"]

    # Merged cell
    sheet.merge_cells("D4:E5")
    merged_cell = sheet["D4"]
    merged_cell.value = "Merged Cell"
    merged_cell.font = Font(name='Verdana', size=14, color='8E7CC3')
    merged_cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')


    workbook.save(file_path)
    print(f"Created failing style sample at: {file_path}")

if __name__ == "__main__":
    create_failing_style_sample("tests/data/failing_style_sample.xlsx")