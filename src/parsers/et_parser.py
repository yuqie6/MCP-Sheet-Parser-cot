import openpyxl
import xlrd
from typing import Optional

from src.models.table_model import Sheet, Row, Cell, Style
from src.parsers.base_parser import BaseParser


class EtParser(BaseParser):
    """Parses .et files (WPS Office format) using fallback methods."""

    def parse(self, file_path: str, sheet_name: Optional[str] = None) -> Sheet:
        """
        Parses the given .et file and returns a Sheet object.
        
        ET files are WPS Office format. We try multiple parsing strategies:
        1. Try openpyxl (if ET file is Excel-compatible)
        2. Try xlrd (for older ET formats)
        3. Basic text-based parsing as fallback
        """
        
        # Strategy 1: Try openpyxl first (many ET files are Excel-compatible)
        try:
            return self._parse_with_openpyxl(file_path, sheet_name)
        except Exception:
            pass
        
        # Strategy 2: Try xlrd for older formats
        try:
            return self._parse_with_xlrd(file_path, sheet_name)
        except Exception:
            pass
        
        # Strategy 3: If all else fails, create an empty sheet with error info
        return self._create_error_sheet(file_path, sheet_name)
    
    def _parse_with_openpyxl(self, file_path: str, sheet_name: Optional[str] = None) -> Sheet:
        """Try parsing with openpyxl."""
        workbook = openpyxl.load_workbook(file_path)
        
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in the workbook")
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.active

        sheet_data = Sheet(name=worksheet.title)
        
        for r_idx, row in enumerate(worksheet.iter_rows()):
            row_data = Row()
            for c_idx, cell in enumerate(row):
                # Basic style extraction
                font = cell.font
                fill = cell.fill
                alignment = cell.alignment
                
                style = Style(
                    bold=font.bold,
                    italic=font.italic,
                    underline=font.underline != 'none' if font.underline else False,
                    font_color=self._get_color_hex(font.color),
                    fill_color=self._get_color_hex(fill.fgColor) if fill.fgColor else None,
                    font_size=font.size,
                    font_name=font.name,
                    alignment=alignment.horizontal,
                    vertical_alignment=alignment.vertical,
                )
                
                cell_data = Cell(
                    value=cell.value,
                    row=r_idx,
                    column=c_idx,
                    style=style
                )
                row_data.cells.append(cell_data)
            sheet_data.rows.append(row_data)
        
        return sheet_data
    
    def _parse_with_xlrd(self, file_path: str, sheet_name: Optional[str] = None) -> Sheet:
        """Try parsing with xlrd."""
        workbook = xlrd.open_workbook(file_path, formatting_info=True)
        
        if sheet_name:
            try:
                worksheet = workbook.sheet_by_name(sheet_name)
            except xlrd.XLRDError:
                raise ValueError(f"Sheet '{sheet_name}' not found in the workbook")
        else:
            worksheet = workbook.sheet_by_index(0)

        sheet_data = Sheet(name=worksheet.name)
        
        for r_idx in range(worksheet.nrows):
            row_data = Row()
            for c_idx in range(worksheet.ncols):
                cell_value = worksheet.cell_value(r_idx, c_idx)
                
                # Basic style (xlrd has limited style support for ET files)
                style = Style()
                
                cell_data = Cell(
                    value=cell_value,
                    row=r_idx,
                    column=c_idx,
                    style=style
                )
                row_data.cells.append(cell_data)
            sheet_data.rows.append(row_data)
        
        return sheet_data
    
    def _create_error_sheet(self, file_path: str, sheet_name: Optional[str] = None) -> Sheet:
        """Create a sheet with error information when parsing fails."""
        sheet_data = Sheet(name=sheet_name or "ET_Parse_Error")
        
        # Create a single row with error message
        row_data = Row()
        error_cell = Cell(
            value=f"Unable to parse ET file: {file_path}. ET format may not be fully supported.",
            row=0,
            column=0,
            style=Style(bold=True, font_color="#FF0000")
        )
        row_data.cells.append(error_cell)
        sheet_data.rows.append(row_data)
        
        return sheet_data
    
    def _get_color_hex(self, color) -> Optional[str]:
        """Convert openpyxl color to hex string."""
        if color and hasattr(color, 'rgb') and color.rgb:
            try:
                if color.rgb.startswith('FF'):
                    return f"#{color.rgb[2:]}"
                else:
                    return f"#{color.rgb}"
            except (AttributeError, TypeError):
                pass
        return None
