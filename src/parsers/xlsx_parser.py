"""
XLSX解析器模块

解析Excel XLSX文件并转换为Sheet对象
包含完整的样式提取、颜色处理、边框识别等功能，支持流式读取。
"""

import logging
import openpyxl
import zipfile
import re
import os
from tempfile import NamedTemporaryFile
import shutil

logger = logging.getLogger(__name__)
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import Cell as OpenpyxlCell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.chart.bar_chart import BarChart
from openpyxl.chart.line_chart import LineChart
from openpyxl.chart.pie_chart import PieChart
from openpyxl.chart.area_chart import AreaChart
from typing import BinaryIO, cast
from collections.abc import Iterator
from io import BytesIO
from src.models.table_model import Sheet, Row, Cell, LazySheet, Chart, ChartPosition
from src.parsers.base_parser import BaseParser
from src.utils.style_parser import extract_style, extract_cell_value
from src.utils.chart_data_extractor import ChartDataExtractor
from src.utils.color_utils import convert_scheme_color_to_hex, generate_pie_color_variants


class XlsxRowProvider:
    """XLSX文件的惰性行提供者，基于openpyxl的read_only流式模式。"""
    
    def __init__(self, file_path: str, sheet_name: str | None = None):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self._total_rows_cache: int | None= None
        self._merged_cells_cache: list[str] | None = None
        self._worksheet_title_cache: str | None = None
    
    def _get_worksheet_info(self):
        """无需读取全部数据即可获取工作表信息。"""
        if self._worksheet_title_cache is None:
            workbook = None
            try:
                workbook = openpyxl.load_workbook(self.file_path, read_only=True)
                worksheet = workbook.active if self.sheet_name is None else workbook[self.sheet_name]
                if worksheet is not None:
                    self._worksheet_title_cache = worksheet.title
                else:
                    self._worksheet_title_cache = ""
            except Exception as e:
                logger.warning(f"Failed to get worksheet info: {e}")
                self._worksheet_title_cache = ""
            finally:
                if workbook is not None:
                    workbook.close()
        return self._worksheet_title_cache
    
    def _get_merged_cells(self) -> list[str]:
        """获取合并单元格信息。"""
        if self._merged_cells_cache is None:
            workbook = None
            try:
                workbook = openpyxl.load_workbook(self.file_path)
                worksheet = workbook.active if self.sheet_name is None else workbook[self.sheet_name]
                if worksheet is not None and hasattr(worksheet, "merged_cells"):
                    self._merged_cells_cache = [str(merged_cell_range) for merged_cell_range in worksheet.merged_cells.ranges]
                else:
                    self._merged_cells_cache = []
            except Exception as e:
                logger.error(f"获取合并单元格信息失败: {e}")
                raise RuntimeError(f"无法加载工作簿以获取合并单元格: {e}") from e
            finally:
                # 确保只有在workbook成功创建后才尝试关闭
                if 'workbook' in locals() and workbook is not None:
                    workbook.close()
        return self._merged_cells_cache
    
    def _parse_row(self, row_cells: tuple) -> Row:
        """将openpyxl单元格元组解析为Row对象。"""
        cells = []
        for cell in row_cells:
            cell_value = cell.value
            cell_style = extract_style(cell) if cell else None
            formula = None
            if isinstance(cell, OpenpyxlCell) and hasattr(cell, 'data_type') and cell.data_type == 'f':
                if hasattr(cell, 'value') and cell.value:
                    formula = str(cell.value)
            cells.append(Cell(value=cell_value, style=cell_style, formula=formula))
        return Row(cells=cells)

    def iter_rows(self, start_row: int = 0, max_rows: int | None = None) -> Iterator[Row]:
        """按需产出完整结构的行。"""
        workbook = None
        try:
            workbook = openpyxl.load_workbook(self.file_path, read_only=True)
            worksheet = workbook.active if self.sheet_name is None else workbook[self.sheet_name]

            if worksheet is not None:
                # 获取工作表的完整尺寸
                max_row = worksheet.max_row or 0
                max_col = worksheet.max_column or 0

                # 计算实际的行范围
                end_row = max_row
                if max_rows is not None:
                    end_row = min(start_row + max_rows, max_row)

                # 使用坐标访问确保完整的行结构
                for row_idx in range(start_row + 1, end_row + 1):
                    cells = []
                    for col_idx in range(1, max_col + 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cells.append(cell)
                    yield self._parse_row(tuple(cells))
        except Exception as e:
            raise RuntimeError(f"流式读取XLSX文件失败: {str(e)}") from e
        finally:
            if workbook is not None:
                workbook.close()
    
    def get_row(self, row_index: int) -> Row:
        """按索引获取完整结构的指定行。"""
        workbook = openpyxl.load_workbook(self.file_path, read_only=True)
        worksheet = workbook.active if self.sheet_name is None else workbook[self.sheet_name]
        try:
            if worksheet is not None:
                max_row = worksheet.max_row or 0
                max_col = worksheet.max_column or 0

                if row_index >= max_row:
                    raise IndexError(f"Row index {row_index} out of range (max: {max_row-1})")

                # 使用坐标访问获取完整的行
                cells = []
                for col_idx in range(1, max_col + 1):
                    cell = worksheet.cell(row=row_index + 1, column=col_idx)
                    cells.append(cell)
                return self._parse_row(tuple(cells))
            raise IndexError(f"Row index {row_index} out of range")
        except IndexError:
            # 重新抛出索引错误
            workbook.close()
            raise
        except Exception as e:
            # 处理其他异常
            workbook.close()
            raise RuntimeError(f"获取XLSX行数据失败: {str(e)}") from e
        finally:
            workbook.close()
    
    def get_total_rows(self) -> int:
        """无需加载全部数据即可获取总行数。"""
        if self._total_rows_cache is None:
            try:
                workbook = openpyxl.load_workbook(self.file_path, read_only=True)
                worksheet = workbook.active if self.sheet_name is None else workbook[self.sheet_name]
                if worksheet is not None and hasattr(worksheet, "max_row"):
                    self._total_rows_cache = worksheet.max_row or 0
                else:
                    self._total_rows_cache = 0
            except Exception as e:
                logger.error(f"获取工作表总行数失败: {e}")
                raise RuntimeError(f"无法加载工作簿以获取总行数: {e}") from e
            finally:
                # 确保只有在workbook成功创建后才尝试关闭
                if 'workbook' in locals() and workbook is not None:
                    workbook.close()
        return self._total_rows_cache


class XlsxParser(BaseParser):
    """
    XLSX文件解析器，支持完整样式提取。

    该解析器处理现代Excel文件（.xlsx），可提取丰富的样式信息，包括字体、颜色、边框和数字格式。
    同时支持通过XlsxRowProvider进行大文件流式读取。
    """

    def __init__(self):
        self.chart_extractor = ChartDataExtractor()

    def parse(self, file_path: str) -> list[Sheet]:
        """
        解析XLSX文件，返回每个工作表对应的Sheet对象列表。
        """
        workbook = None
        data_only_workbook = None

        # 尝试多种加载方式
        load_attempts = [
            {"data_only": False, "keep_vba": False, "keep_links": False},
            {"data_only": True, "keep_vba": False, "keep_links": False},
            {"data_only": True, "keep_vba": False, "keep_links": False, "read_only": True},
        ]

        last_error = None
        for i, kwargs in enumerate(load_attempts):
            try:
                logger.info(f"尝试加载方式 {i+1}: {kwargs}")
                workbook = openpyxl.load_workbook(file_path, **kwargs)

                # 如果成功加载，尝试获取data_only版本
                if not kwargs.get("data_only", False):
                    try:
                        data_only_workbook = openpyxl.load_workbook(file_path, data_only=True, keep_vba=False, keep_links=False)
                    except Exception as e:
                        logger.debug(f"无法加载data_only版本，使用原始工作簿: {e}")
                        data_only_workbook = workbook
                else:
                    data_only_workbook = workbook

                logger.info(f"成功使用方式 {i+1} 加载文件")
                break

            except Exception as e:
                last_error = e
                logger.warning(f"加载方式 {i+1} 失败: {e}")
                continue

        if workbook is None:
            # 检查是否是样式兼容性问题
            if last_error and "Fill" in str(last_error):
                logger.warning(f"检测到样式兼容性问题，尝试修复Excel文件: {last_error}")
                try:
                    fixed_file = self._fix_excel_styles(file_path)
                    logger.info(f"样式修复成功，重新尝试解析: {fixed_file}")

                    # 使用修复后的文件重新尝试解析
                    for i, kwargs in enumerate(load_attempts):
                        try:
                            workbook = openpyxl.load_workbook(fixed_file, **kwargs)
                            if not kwargs.get("data_only", False):
                                try:
                                    data_only_workbook = openpyxl.load_workbook(fixed_file, data_only=True, keep_vba=False, keep_links=False)
                                except Exception as e:
                                    logger.debug(f"修复文件无法加载data_only版本，使用原始工作簿: {e}")
                                    data_only_workbook = workbook
                            else:
                                data_only_workbook = workbook
                            logger.info(f"修复后文件解析成功")
                            break
                        except Exception as e:
                            continue

                    if workbook is None:
                        logger.error("修复后的文件仍然无法解析")

                except Exception as fix_error:
                    logger.error(f"样式修复失败: {fix_error}")

            if workbook is None:
                logger.error(f"openpyxl所有加载方式都失败，尝试使用XLS解析器作为备选")
                try:
                    # 尝试使用XLS解析器作为备选
                    from .xls_parser import XlsParser
                    xls_parser = XlsParser()
                    logger.info("使用XLS解析器作为备选方案")
                    return xls_parser.parse(file_path)
                except Exception as xls_error:
                    logger.error(f"XLS解析器也失败: {xls_error}")
                    raise IOError(f"无法加载Excel文件 (openpyxl: {last_error}, xlrd: {xls_error})")

        sheets = []
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            data_only_worksheet = data_only_workbook[sheet_name] if data_only_workbook else worksheet

            sheet = self._parse_sheet(worksheet, data_only_worksheet)
            sheets.append(sheet)
            
        return sheets

    def _parse_sheet(self, worksheet: Worksheet, data_only_worksheet: Worksheet) -> Sheet:
        """解析单个工作表的辅助方法。"""
        max_row = worksheet.max_row or 0
        max_col = worksheet.max_column or 0

        rows = []
        for row_idx in range(1, max_row + 1):
            cells = []
            for col_idx in range(1, max_col + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                data_cell = data_only_worksheet.cell(row=row_idx, column=col_idx)
                
                cell_style = extract_style(cell)
                
                if cell.data_type == 'f' and cell.value:
                    cell_value = data_cell.value
                    formula = str(cell.value)
                else:
                    cell_value = extract_cell_value(cell)
                    formula = None
                
                parsed_cell = Cell(
                    value=cell_value,
                    style=cell_style,
                    formula=formula
                )
                cells.append(parsed_cell)
            rows.append(Row(cells=cells))

        merged_cells = [str(merged_cell_range) for merged_cell_range in worksheet.merged_cells.ranges]
        
        column_widths = {col_idx - 1: worksheet.column_dimensions[get_column_letter(col_idx)].width 
                         for col_idx in range(1, max_col + 1) if worksheet.column_dimensions[get_column_letter(col_idx)].width}
                         
        row_heights = {row_idx - 1: worksheet.row_dimensions[row_idx].height 
                       for row_idx in range(1, max_row + 1) if worksheet.row_dimensions[row_idx].height}

        default_col_width = worksheet.sheet_format.defaultColWidth or 8.43
        default_row_height = worksheet.sheet_format.defaultRowHeight or 18.0
        
        charts = self._extract_charts(worksheet)
        images = self._extract_images(worksheet)
        all_visuals = charts + images

        return Sheet(
            name=worksheet.title,
            rows=rows,
            merged_cells=merged_cells,
            charts=all_visuals,
            column_widths=column_widths,
            row_heights=row_heights,
            default_column_width=default_col_width,
            default_row_height=default_row_height
        )

    def _extract_images(self, worksheet: Worksheet) -> list[Chart]:
        """提取工作表中的嵌入图片。"""
        images = []
        worksheet_images = getattr(worksheet, '_images', [])
        for image in worksheet_images:
            if isinstance(image, OpenpyxlImage):
                try:
                    img_data = None

                    # 方法1：尝试从ref（BytesIO对象）读取
                    if hasattr(image, 'ref') and image.ref:
                        ref = image.ref
                        try:
                            # 检查是否是类文件对象（有read和seek方法）
                            if hasattr(ref, 'read') and hasattr(ref, 'seek') and callable(getattr(ref, 'read', None)) and callable(getattr(ref, 'seek', None)):
                                # ref是文件对象（如BytesIO），使用类型转换来避免类型检查错误
                                file_like_ref = cast(BinaryIO, ref)
                                file_like_ref.seek(0)  # 确保从开头读取
                                img_data = file_like_ref.read()
                            elif isinstance(ref, bytes):
                                img_data = ref
                        except (AttributeError, OSError, IOError, TypeError):
                            # 如果seek或read操作失败，继续尝试其他方法
                            pass

                    # 方法2：尝试从_data属性或方法获取
                    if not img_data and hasattr(image, '_data'):
                        try:
                            data_attr = getattr(image, '_data', None)
                            if callable(data_attr):
                                img_data = data_attr()
                            elif data_attr and not callable(data_attr):
                                img_data = data_attr
                        except Exception:
                            pass

                    anchor_str = "A1"  # 默认位置
                    position = None

                    if hasattr(image, 'anchor') and image.anchor:
                        try:
                            # 安全地处理anchor对象
                            anchor = image.anchor
                            if hasattr(anchor, '_from') and getattr(anchor, '_from', None):
                                # TwoCellAnchor类型 - 提取详细位置信息
                                from_cell = getattr(anchor, '_from', None)
                                to_cell = getattr(anchor, '_to', None)

                                if from_cell and hasattr(from_cell, 'col') and hasattr(from_cell, 'row'):
                                    # 转换为Excel单元格引用
                                    from openpyxl.utils import get_column_letter
                                    col_letter = get_column_letter(from_cell.col + 1)  # openpyxl使用0基索引
                                    anchor_str = f"{col_letter}{from_cell.row + 1}"

                                    # 创建位置对象
                                    from_col_offset = getattr(from_cell, 'colOff', 0) if hasattr(from_cell, 'colOff') else 0
                                    from_row_offset = getattr(from_cell, 'rowOff', 0) if hasattr(from_cell, 'rowOff') else 0

                                    # 如果有to_cell，使用它；否则估算图片大小
                                    if to_cell and hasattr(to_cell, 'col') and hasattr(to_cell, 'row'):
                                        to_col = to_cell.col
                                        to_row = to_cell.row
                                        to_col_offset = getattr(to_cell, 'colOff', 0) if hasattr(to_cell, 'colOff') else 0
                                        to_row_offset = getattr(to_cell, 'rowOff', 0) if hasattr(to_cell, 'rowOff') else 0
                                    else:
                                        # 估算图片占用2列3行
                                        to_col = from_cell.col + 2
                                        to_row = from_cell.row + 3
                                        to_col_offset = 0
                                        to_row_offset = 0

                                    position = ChartPosition(
                                        from_col=from_cell.col,
                                        from_row=from_cell.row,
                                        from_col_offset=from_col_offset,
                                        from_row_offset=from_row_offset,
                                        to_col=to_col,
                                        to_row=to_row,
                                        to_col_offset=to_col_offset,
                                        to_row_offset=to_row_offset
                                    )
                            else:
                                # 其他类型的anchor，转换为字符串
                                anchor_str = str(anchor)
                        except Exception:
                            anchor_str = "A1"

                    image_chart = Chart(
                        name=f"Image {len(images) + 1}",
                        type="image",
                        anchor=anchor_str,
                        position=position  # 添加位置信息
                    )
                    image_chart.chart_data = {'image_data': img_data, 'type': 'image'}
                    images.append(image_chart)
                except Exception as e:
                    import logging
                    import traceback
                    logger = logging.getLogger(__name__)
                    logger.error(f"Failed to extract image: {e}")
                    logger.error(f"Traceback: {traceback.format_exc()}")
                    # 即使提取失败，也创建一个占位符图片对象
                    try:
                        placeholder_chart = Chart(
                            name=f"Image {len(images) + 1} (Failed)",
                            type="image",
                            anchor="A1"
                        )
                        placeholder_chart.chart_data = {'image_data': None, 'type': 'image', 'error': str(e)}
                        images.append(placeholder_chart)
                        logger.info(f"Created placeholder for failed image extraction")
                    except Exception as e2:
                        logger.error(f"Failed to create placeholder: {e2}")
        return images


    
    def _extract_charts(self, worksheet) -> list[Chart]:
        """提取工作表中的图表并保存原始数据。"""
        charts = []
        if not hasattr(worksheet, '_charts'):
            return charts

        for chart_drawing in worksheet._charts:
            chart_type = "unknown"
            if isinstance(chart_drawing, BarChart):
                chart_type = 'bar'
            elif isinstance(chart_drawing, LineChart):
                chart_type = 'line'
            elif isinstance(chart_drawing, PieChart):
                chart_type = 'pie'
            elif isinstance(chart_drawing, AreaChart):
                chart_type = 'area'

            try:
                # 增强：直接从图表的XML中解析颜色
                chart_data = self._extract_chart_data(chart_drawing, chart_type)
                
                
                # Safely get chart title using the extractor
                chart_title = None
                if chart_drawing.title:
                    chart_title = self.chart_extractor.extract_axis_title(chart_drawing.title)
                
                # If extraction fails or returns None, provide a meaningful fallback
                if not chart_title:
                    chart_title = f"Chart {len(charts) + 1}"
                
                # 提取详细的定位信息
                anchor_value = "A1"  # 默认位置
                position = None
                
                if hasattr(chart_drawing, 'anchor') and chart_drawing.anchor:
                    try:
                        anchor = chart_drawing.anchor
                        if hasattr(anchor, '_from') and getattr(anchor, '_from', None) and hasattr(anchor, 'to') and getattr(anchor, 'to', None):
                            # 提取完整定位数据
                            from_cell = getattr(anchor, '_from', None)
                            to_cell = getattr(anchor, 'to', None)
                            
                            if (from_cell and to_cell and 
                                hasattr(from_cell, 'col') and hasattr(from_cell, 'row') and
                                hasattr(to_cell, 'col') and hasattr(to_cell, 'row')):
                                
                                from_col = from_cell.col
                                from_row = from_cell.row
                                
                                # 使用原始EMU值，而不是转换后的像素
                                position = ChartPosition(
                                    from_col=from_col,
                                    from_row=from_row,
                                    from_col_offset=getattr(from_cell, 'colOff', 0),
                                    from_row_offset=getattr(from_cell, 'rowOff', 0),
                                    to_col=to_cell.col,
                                    to_row=to_cell.row,  
                                    to_col_offset=getattr(to_cell, 'colOff', 0),
                                    to_row_offset=getattr(to_cell, 'rowOff', 0)
                                )
                                
                                # 生成简单的anchor字符串用于向后兼容
                                from openpyxl.utils import get_column_letter
                                col_letter = get_column_letter(from_col + 1)
                                anchor_value = f"{col_letter}{from_row + 1}"
                                
                        elif hasattr(anchor, 'cell'):
                            # 其他类型的anchor
                            anchor_value = str(getattr(anchor, 'cell', 'A1'))
                    except Exception as e:
                        import logging
                        logger = logging.getLogger(__name__)
                        logger.warning(f"Failed to extract positioning for chart {chart_title}: {e}")
                        anchor_value = "A1"
                        
                charts.append(Chart(
                    name=chart_title,
                    type=chart_type,
                    anchor=anchor_value,
                    chart_data=chart_data,
                    position=position
                ))
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Failed to extract chart {chart_drawing.title}: {e}")

        return charts


    
    def supports_streaming(self) -> bool:
        """XLSX解析器支持流式处理。"""
        return True
    
    def create_lazy_sheet(self, file_path: str, sheet_name: str | None = None) -> LazySheet:
        """
        创建用于流式读取XLSX数据的LazySheet。

        参数：
            file_path: XLSX文件的绝对路径。
            sheet_name: 要解析的工作表名称（可选）。

        返回：
            可按需流式读取数据的LazySheet对象。
        """
        provider = XlsxRowProvider(file_path, sheet_name)
        name = provider._get_worksheet_info()
        merged_cells = provider._get_merged_cells()
        return LazySheet(name=name, provider=provider, merged_cells=merged_cells)

    def _extract_chart_data(self, chart, chart_type: str) -> dict:
        """
        提取图表的原始数据，用于SVG渲染。
        
        参数：
            chart: openpyxl图表对象
            chart_type: 图表类型
            
        返回：
            包含图表数据的字典
        """
        chart_data = {
            'type': chart_type,
            'title': self.chart_extractor.extract_axis_title(chart.title) if chart.title else '',
            'series': [],
            'x_axis_title': '',
            'y_axis_title': '',
            'position': {},  # 添加位置信息
            'size': {},      # 添加尺寸信息
            'colors': [],    # 添加原始颜色信息
            'legend': {},    # 添加图例信息
            'annotations': [],  # 添加注释信息
            'data_labels': {}   # 添加数据标签信息
        }
        
        # 提取位置和尺寸信息
        if hasattr(chart, 'anchor') and chart.anchor:
            try:
                anchor = chart.anchor
                if hasattr(anchor, '_from') and getattr(anchor, '_from', None):
                    from_cell = getattr(anchor, '_from', None)
                    if from_cell:
                        chart_data['position']['from_col'] = getattr(from_cell, 'col', 0)
                        chart_data['position']['from_row'] = getattr(from_cell, 'row', 0)
                        chart_data['position']['from_col_offset'] = getattr(from_cell, 'colOff', 0)
                        chart_data['position']['from_row_offset'] = getattr(from_cell, 'rowOff', 0)
                
                if hasattr(anchor, 'to') and getattr(anchor, 'to', None):
                    to_cell = getattr(anchor, 'to', None)
                    if to_cell:
                        chart_data['position']['to_col'] = getattr(to_cell, 'col', 0)
                        chart_data['position']['to_row'] = getattr(to_cell, 'row', 0)
                        chart_data['position']['to_col_offset'] = getattr(to_cell, 'colOff', 0)
                        chart_data['position']['to_row_offset'] = getattr(to_cell, 'rowOff', 0)
                
                if hasattr(anchor, 'ext') and anchor.ext:
                    ext = anchor.ext
                    chart_data['size']['width_emu'] = ext.cx
                    chart_data['size']['height_emu'] = ext.cy
                    # EMU转像素 (1 EMU = 1/914400 inch, 1 inch = 96 px)
                    chart_data['size']['width_px'] = int(ext.cx / 914400 * 96)
                    chart_data['size']['height_px'] = int(ext.cy / 914400 * 96)
            except Exception:
                pass
        
        # 提取轴标题
        try:
            if chart.x_axis and chart.x_axis.title:
                x_title = self.chart_extractor.extract_axis_title(chart.x_axis.title)
                if x_title:
                    chart_data['x_axis_title'] = x_title
            if chart.y_axis and chart.y_axis.title:
                y_title = self.chart_extractor.extract_axis_title(chart.y_axis.title)
                if y_title:
                    chart_data['y_axis_title'] = y_title
        except Exception:
            pass
        
        # 提取轴缩放
        try:
            if hasattr(chart, 'y_axis') and chart.y_axis and hasattr(chart.y_axis, 'scaling'):
                scaling = chart.y_axis.scaling
                if hasattr(scaling, 'min') and scaling.min is not None:
                    chart_data['y_axis_min'] = float(scaling.min)
                if hasattr(scaling, 'max') and scaling.max is not None:
                    chart_data['y_axis_max'] = float(scaling.max)
            if hasattr(chart, 'x_axis') and chart.x_axis and hasattr(chart.x_axis, 'scaling'):
                scaling = chart.x_axis.scaling
                if hasattr(scaling, 'min') and scaling.min is not None:
                    chart_data['x_axis_min'] = float(scaling.min)
                if hasattr(scaling, 'max') and scaling.max is not None:
                    chart_data['x_axis_max'] = float(scaling.max)
        except (ValueError, TypeError, AttributeError):
            pass 
        
        # 提取系列数据
        if isinstance(chart, (BarChart, LineChart, AreaChart)):
            for series in chart.series:
                # 使用chart_extractor来提取系列名称
                series_name = None
                if series.tx:
                    series_name = self.chart_extractor.extract_axis_title(series.tx)
                if not series_name:
                    series_name = f"Series {len(chart_data['series']) + 1}"

                series_data = {
                    'name': series_name,
                    'x_data': [],
                    'y_data': [],
                    'color': None  # 添加颜色信息
                }
                
                # 提取系列颜色
                series_color = self.chart_extractor.extract_series_color(series)
                if series_color:
                    series_data['color'] = series_color
                    chart_data['colors'].append(series_color)

                # 提取数据标签信息
                data_labels = self.chart_extractor.extract_data_labels(series)
                if data_labels['enabled']:
                    series_data['data_labels'] = data_labels
                
                # 获取y轴数据
                y_data = self.chart_extractor.extract_series_y_data(series)
                if y_data:
                    series_data['y_data'] = y_data
                
                # 获取x轴数据
                x_data = self.chart_extractor.extract_series_x_data(series)
                if x_data:
                    series_data['x_data'] = x_data
                elif y_data:
                    # 如果没有x轴数据，生成默认标签
                    series_data['x_data'] = [f"Item {i+1}" for i in range(len(y_data))]
                
                if series_data['y_data']:
                    chart_data['series'].append(series_data)
                    
        elif isinstance(chart, PieChart):
            if chart.series and len(chart.series) > 0:
                series = chart.series[0]
                # 使用chart_extractor来提取系列名称
                series_name = None
                if series.tx:
                    series_name = self.chart_extractor.extract_axis_title(series.tx)
                if not series_name:
                    series_name = "Pie Series"

                series_data = {
                    'name': series_name,
                    'x_data': [],  # 标签
                    'y_data': [],   # 数值
                    'colors': []   # 饼图每个片段的颜色
                }
                
                # 获取标签数据
                x_data = self.chart_extractor.extract_series_x_data(series)
                if x_data:
                    series_data['x_data'] = x_data
                
                # 获取数值数据
                y_data = self.chart_extractor.extract_series_y_data(series)
                if y_data:
                    series_data['y_data'] = y_data
                
                # 提取饼图颜色（每个数据点可能有不同颜色）
                pie_colors = self.chart_extractor.extract_pie_chart_colors(series)
                if pie_colors:
                    series_data['colors'] = pie_colors
                    chart_data['colors'] = pie_colors

                # 提取数据标签信息
                data_labels = self.chart_extractor.extract_data_labels(series)
                if data_labels['enabled']:
                    series_data['data_labels'] = data_labels
                
                # 如果没有标签，生成默认标签
                if not series_data['x_data'] and series_data['y_data']:
                    series_data['x_data'] = [f"Item {i+1}" for i in range(len(series_data['y_data']))]
                
                if series_data['y_data']:
                    chart_data['series'].append(series_data)

        # 提取图例信息
        legend_info = self.chart_extractor.extract_legend_info(chart)
        if legend_info['enabled']:
            # 如果图例条目没有文本，根据图表类型使用不同的策略
            if isinstance(chart, PieChart) and chart_data['series']:
                # 对于饼图，图例条目应该对应每个片段的标签
                pie_series = chart_data['series'][0]
                x_data = pie_series.get('x_data', [])
                for i, entry in enumerate(legend_info.get('entries', [])):
                    if not entry.get('text'):
                        if i < len(x_data):
                            entry['text'] = x_data[i]
                        else:
                            entry['text'] = f"Item {i + 1}"
            else:
                # 对于其他图表类型，使用系列名称
                for i, entry in enumerate(legend_info.get('entries', [])):
                    if not entry.get('text') and i < len(chart_data['series']):
                        entry['text'] = chart_data['series'][i]['name']
            chart_data['legend'] = legend_info

        # 提取注释信息
        annotations = self.chart_extractor.extract_chart_annotations(chart)
        if annotations:
            chart_data['annotations'] = annotations

        return chart_data

    def _fix_excel_styles(self, file_path: str) -> str:
        """
        修复Excel文件中的样式兼容性问题

        参数:
            file_path: 原始Excel文件路径

        返回:
            修复后的Excel文件路径
        """
        # 生成修复后的文件名
        name, ext = os.path.splitext(file_path)
        fixed_file = f"{name}_fixed{ext}"

        logger.info(f"正在修复Excel样式兼容性问题: {file_path} -> {fixed_file}")

        with NamedTemporaryFile(delete=False) as tmp:
            tmp_name = tmp.name

        try:
            zin = zipfile.ZipFile(file_path, "r")
            zout = zipfile.ZipFile(tmp_name, "w")

            for item in zin.infolist():
                buffer = zin.read(item.filename)

                if item.filename == "xl/styles.xml":
                    logger.debug("修复 xl/styles.xml 文件...")
                    styles = buffer.decode("utf-8")

                    # 1. 移除问题的独立空标签
                    styles = styles.replace("<x:fill />", "")

                    # 2. 修复fills部分中的空fill标签
                    pattern = re.compile(r'(<fills count="\d+">)(.*?)(</fills>)', re.DOTALL)

                    def fix_fills_section(match):
                        start_tag = match.group(1)
                        content = match.group(2)
                        end_tag = match.group(3)

                        # 移除空的fill标签
                        content = re.sub(r'<fill\s*/>', '', content)
                        content = re.sub(r'<fill></fill>', '', content)
                        content = re.sub(r'<fill>\s*</fill>', '', content)

                        # 如果内容为空或只有空白，添加一个默认的fill
                        if not content.strip():
                            content = '<fill><patternFill patternType="none"/></fill>'

                        return start_tag + content + end_tag

                    styles = pattern.sub(fix_fills_section, styles)

                    # 3. 额外的清理：移除任何剩余的空fill标签
                    styles = re.sub(r'<fill\s*/>', '', styles)
                    styles = re.sub(r'<fill></fill>', '', styles)
                    styles = re.sub(r'<fill>\s*</fill>', '', styles)

                    # 4. 确保fills部分有正确的count属性
                    fills_pattern = re.compile(r'<fills count="(\d+)">(.*?)</fills>', re.DOTALL)

                    def fix_fills_count(match):
                        old_count = match.group(1)
                        content = match.group(2)

                        # 计算实际的fill元素数量
                        actual_count = len(re.findall(r'<fill[^>]*>', content))

                        if int(old_count) != actual_count:
                            logger.debug(f"修正fills count: {old_count} -> {actual_count}")
                            return f'<fills count="{actual_count}">{content}</fills>'

                        return match.group(0)

                    styles = fills_pattern.sub(fix_fills_count, styles)
                    buffer = styles.encode("utf-8")

                zout.writestr(item, buffer)

            zin.close()
            zout.close()

            # 将临时文件移动到目标位置
            shutil.move(tmp_name, fixed_file)
            logger.info(f"Excel样式修复完成: {fixed_file}")

            return fixed_file

        except Exception as e:
            logger.error(f"Excel样式修复失败: {e}")
            # 清理临时文件
            if os.path.exists(tmp_name):
                os.unlink(tmp_name)
            raise
