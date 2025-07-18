"""
XLSM格式解析器模块

基于openpyxl库实现XLSM格式（Excel宏文件）的解析。
XLSM本质上是带宏的XLSX，样式提取与XLSX完全相同。
"""

import logging
import openpyxl
from src.models.table_model import Sheet, Row, Cell, LazySheet
from src.parsers.xlsx_parser import XlsxParser
from src.utils.style_parser import extract_style

logger = logging.getLogger(__name__)


class XlsmParser(XlsxParser):
    """
    XLSM格式解析器，继承自XlsxParser。

    XLSM格式与XLSX在数据和样式方面完全一致，主要区别是包含VBA宏代码。
    本解析器保留宏信息但不执行，仅专注于数据和样式提取。
    """
    
    def parse(self, file_path: str) -> list[Sheet]:
        """
        解析XLSM文件并返回Sheet对象列表。

        参数:
            file_path: XLSM文件路径

        返回:
            包含完整数据和样式的Sheet对象列表

        异常:
            RuntimeError: 当解析失败时
        """
        try:
            # 使用keep_vba=True保留宏信息，但不执行
            workbook = openpyxl.load_workbook(file_path, keep_vba=True)

            if not workbook.worksheets:
                raise ValueError("工作簿不包含任何工作表")

            # 记录宏文件信息（用于调试和日志）
            self._log_macro_info(workbook, file_path)

            sheets = []

            # 解析所有工作表
            for worksheet in workbook.worksheets:
                # 解析数据和样式（使用与XlsxParser相同的完整性逻辑）
                # 获取工作表的实际尺寸，确保包含所有数据和样式
                max_row = worksheet.max_row or 0
                max_col = worksheet.max_column or 0

                rows = []
                # 使用坐标访问方式确保完整的表格结构
                for row_idx in range(1, max_row + 1):
                    cells = []
                    for col_idx in range(1, max_col + 1):
                        # 直接通过坐标访问单元格，确保包含空单元格
                        cell = worksheet.cell(row=row_idx, column=col_idx)

                        # 提取单元格值和样式
                        cell_value = cell.value
                        cell_style = extract_style(cell)

                        # 创建Cell对象
                        parsed_cell = Cell(
                            value=cell_value,
                            style=cell_style
                        )
                        cells.append(parsed_cell)
                    rows.append(Row(cells=cells))

                # 提取合并单元格信息
                merged_cells = [str(merged_cell_range) for merged_cell_range in worksheet.merged_cells.ranges]

                sheet = Sheet(
                    name=worksheet.title,
                    rows=rows,
                    merged_cells=merged_cells
                )
                sheets.append(sheet)

            return sheets
            
        except Exception as e:
            logger.error(f"解析XLSM文件失败: {e}")
            raise RuntimeError(f"无法解析XLSM文件 {file_path}: {str(e)}")
    
    def _log_macro_info(self, workbook, file_path: str) -> None:
        """
        记录宏文件的相关信息，用于调试和日志。
        
        参数:
            workbook: openpyxl工作簿对象
            file_path: 文件路径
        """
        try:
            # 检查是否包含VBA项目
            has_vba = hasattr(workbook, 'vba_archive') and workbook.vba_archive is not None
            
            if has_vba:
                logger.info(f"XLSM文件 {file_path} 包含VBA宏代码，已保留但不执行")
                
                # 尝试获取VBA项目的基本信息
                try:
                    vba_archive = workbook.vba_archive
                    if hasattr(vba_archive, 'namelist'):
                        vba_files = vba_archive.namelist()
                        logger.debug(f"VBA文件数量: {len(vba_files)}")
                        
                        # 记录主要的VBA组件
                        vba_modules = [f for f in vba_files if f.endswith('.bas') or f.endswith('.cls') or f.endswith('.frm')]
                        if vba_modules:
                            logger.debug(f"VBA模块: {vba_modules}")
                            
                except Exception as vba_e:
                    logger.debug(f"获取VBA详细信息失败: {vba_e}")
            else:
                logger.info(f"XLSM文件 {file_path} 不包含VBA宏代码")
                
        except Exception as e:
            logger.debug(f"记录宏信息失败: {e}")
    
    def get_macro_info(self, file_path: str) -> dict:
        """
        获取XLSM文件的宏信息摘要。
        
        参数:
            file_path: XLSM文件路径

        返回:
            包含宏信息的字典
        """
        macro_info = {
            "has_macros": False,
            "vba_files_count": 0,
            "vba_modules": [],
            "file_path": file_path
        }
        
        try:
            workbook = openpyxl.load_workbook(file_path, keep_vba=True)
            
            # 检查VBA存在性
            if hasattr(workbook, 'vba_archive') and workbook.vba_archive is not None:
                macro_info["has_macros"] = True
                
                try:
                    vba_archive = workbook.vba_archive
                    if hasattr(vba_archive, 'namelist'):
                        vba_files = vba_archive.namelist()
                        macro_info["vba_files_count"] = len(vba_files)
                        
                        # 提取VBA模块名称
                        vba_modules = [f for f in vba_files if f.endswith('.bas') or f.endswith('.cls') or f.endswith('.frm')]
                        macro_info["vba_modules"] = vba_modules
                        
                except Exception as vba_e:
                    logger.debug(f"获取VBA详细信息失败: {vba_e}")
            
            workbook.close()
            
        except Exception as e:
            logger.warning(f"获取宏信息失败: {e}")
        
        return macro_info
    
    def supports_streaming(self) -> bool:
        """XLSM解析器支持流式处理（继承自XLSX）。"""
        return True

    def create_lazy_sheet(self, file_path: str, sheet_name: str | None = None) -> LazySheet:
        """
        为XLSM创建可按需流式读取的惰性表对象。
        XLSM文件的流式处理方式与XLSX完全一致。

        参数：
            file_path: XLSM文件路径
            sheet_name: 工作表名称（可选）

        返回：
            LazySheet对象
        """
        # 复用XLSX的流式实现
        from .xlsx_parser import XlsxRowProvider

        provider = XlsxRowProvider(file_path, sheet_name)
        name = provider._get_worksheet_info()
        merged_cells = provider._get_merged_cells()
        return LazySheet(name=name, provider=provider, merged_cells=merged_cells)
    
    def is_macro_enabled_file(self, file_path: str) -> bool:
        """
        检查文件是否为启用宏的Excel文件。
        
        参数:
            file_path: 文件路径

        返回:
            如果文件包含宏则返回True，否则返回False
        """
        try:
            # 首先检查文件扩展名
            if not file_path.lower().endswith('.xlsm'):
                return False
            
            # 然后检查实际的宏内容
            macro_info = self.get_macro_info(file_path)
            return macro_info["has_macros"]
            
        except Exception as e:
            logger.debug(f"检查宏状态失败: {e}")
            return False
