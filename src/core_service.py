"""
核心服务模块 - 简洁版

提供3个核心功能的业务逻辑实现：
1. 获取表格元数据
2. 解析表格数据为JSON
3. 将修改应用回文件
"""

import logging
from datetime import date, datetime
from pathlib import Path
from typing import Any

from .utils.range_parser import parse_range_string
from .utils.style_parser import style_to_dict
from .parsers.factory import ParserFactory
from .models.table_model import Sheet
from .converters.html_converter import HTMLConverter
from .streaming import StreamingTableReader, ChunkFilter
from .unified_config import get_config
from .cache import get_cache_manager
from .exceptions import FileNotFoundError
from .validators import validate_file_input

logger = logging.getLogger(__name__)


class CoreService:
    """核心服务类，提供表格处理的核心功能。"""
    
    def __init__(self):
        self.parser_factory = ParserFactory()
    

    def parse_sheet(self, file_path: str, sheet_name: str | None = None,
                   range_string: str | None = None,
                   enable_streaming: bool = True,
                   streaming_threshold: int | None = None) -> dict[str, Any]:
        """
        解析表格文件为标准化的JSON格式。

        参数:
            file_path: 文件路径
            sheet_name: 工作表名称（可选）
            range_string: 单元格范围（可选）
            enable_streaming: 是否启用自动流式读取（默认True）
            streaming_threshold: 流式读取的单元格数量阈值，None时使用配置默认值

        返回:
            标准化的TableModel JSON
        """
        # 获取当前配置
        current_config = get_config()
        if streaming_threshold is None:
            streaming_threshold = current_config.streaming_threshold_cells
        try:
            # 验证文件输入
            validated_path, _ = validate_file_input(file_path)
            
            # 尝试从缓存获取数据
            cache_manager = get_cache_manager()
            cached_data = cache_manager.get(file_path, range_string, sheet_name)
            if cached_data is not None:
                logger.info(f"从缓存获取数据: {file_path}")
                return cached_data['data']
            
            # 获取解析器
            parser = self.parser_factory.get_parser(str(validated_path))

            # 检查是否应该使用流式读取
            if enable_streaming and self._should_use_streaming(str(validated_path), streaming_threshold):
                json_data = self._parse_sheet_streaming(str(validated_path), sheet_name, range_string)
            else:
                # 使用传统方法
                sheets = parser.parse(str(validated_path))
                
                # 如果指定了工作表名称，则选择对应的工作表
                if sheet_name:
                    target_sheet = next((s for s in sheets if s.name == sheet_name), None)
                    if not target_sheet:
                        raise ValueError(f"工作表 '{sheet_name}' 不存在。")
                # 否则，默认使用第一个工作表
                else:
                    if not sheets:
                        raise ValueError("文件中没有找到任何工作表。")
                    target_sheet = sheets[0]

                # 检查工作表是否为空
                if not target_sheet:
                    logger.warning("目标工作表为None")
                    return {
                        "sheet_name": "Empty",
                        "headers": [],
                        "rows": [],
                        "total_rows": 0,
                        "total_columns": 0,
                        "size_info": {
                            "total_cells": 0,
                            "processing_mode": "empty",
                            "recommendation": "工作表为空，无数据可显示"
                        }
                    }

                if not target_sheet.rows:
                    logger.warning(f"工作表 '{target_sheet.name}' 为空")
                    return {
                        "sheet_name": target_sheet.name,
                        "headers": [],
                        "rows": [],
                        "total_rows": 0,
                        "total_columns": 0,
                        "size_info": {
                            "total_cells": 0,
                            "processing_mode": "empty",
                            "recommendation": "工作表为空，无数据可显示"
                        }
                    }
                
                # 转换为标准化JSON格式
                json_data = self._sheet_to_json(target_sheet, range_string)
            
            # 缓存解析结果
            cache_manager.set(file_path, json_data, range_string, sheet_name)
            logger.debug(f"数据已缓存: {file_path}")
            
            return json_data
            
        except Exception as e:
            logger.error(f"解析表格失败: {e}")
            raise

    def parse_sheet_optimized(self, file_path: str, sheet_name: str | None = None,
                             range_string: str | None = None, include_full_data: bool = False,
                             include_styles: bool = False, preview_rows: int = 5,
                             max_rows: int | None = None) -> dict[str, Any]:
        """
        参数：
            file_path: 文件路径
            sheet_name: 工作表名称（可选）
            range_string: 单元格范围（可选）
            include_full_data: 是否返回完整数据（默认False，只返回概览）
            include_styles: 是否包含样式信息（默认False）
            preview_rows: 预览行数（默认5行）
            max_rows: 最大返回行数（可选）

        返回：
            优化后的JSON数据
        """
        try:
            # 验证文件输入
            validated_path, _ = validate_file_input(file_path)

            # 获取解析器
            parser = self.parser_factory.get_parser(str(validated_path))

            # 解析文件
            sheets = parser.parse(str(validated_path))

            # 选择目标工作表
            if sheet_name:
                target_sheet = next((s for s in sheets if s.name == sheet_name), None)
                if not target_sheet:
                    available_sheets = [s.name for s in sheets]
                    raise ValueError(f"工作表 '{sheet_name}' 不存在。可用工作表: {available_sheets}")
            else:
                if not sheets:
                    raise ValueError("文件中没有找到任何工作表。")
                target_sheet = sheets[0]

            # 处理范围选择
            if range_string:
                try:
                    start_row, start_col, end_row, end_col = parse_range_string(range_string)
                    return self._extract_range_data(target_sheet, start_row, start_col, end_row, end_col, include_styles)
                except ValueError as e:
                    raise ValueError(f"范围格式错误: {e}")

            # 根据参数返回不同级别的数据
            return self._extract_optimized_data(
                target_sheet,
                include_full_data=include_full_data,
                include_styles=include_styles,
                preview_rows=preview_rows,
                max_rows=max_rows
            )

        except Exception as e:
            logger.error(f"优化解析失败: {e}")
            raise

    def convert_to_html(self, file_path: str, output_path: str | None = None,
                       sheet_name: str | None = None,
                       page_size: int | None = None, page_number: int | None = None,
                       header_rows: int = 1) -> list[dict[str, Any]]:
        """
        将表格文件转换为HTML文件。

        参数：
            file_path: 源文件路径
            output_path: 输出HTML文件路径，如果为None则生成默认路径
            page_size: 分页大小（每页行数），如果为None则不分页
            page_number: 页码（从1开始），如果为None则显示第1页
            header_rows: 表头行数，默认第一行为表头

        返回：
            转换结果信息
        """
        try:
            # 验证文件存在
            path = Path(file_path)
            if not path.exists():
                raise FileNotFoundError(f"文件不存在: {file_path}")

            # 生成默认输出路径
            if output_path is None:
                output_path = str(path.with_suffix('.html'))

            # 获取解析器并解析
            parser = self.parser_factory.get_parser(file_path)
            sheets: list[Sheet] = parser.parse(file_path)

            # Filter sheets if a specific sheet_name is provided
            sheets_to_convert = sheets
            if sheet_name:
                sheets_to_convert = [s for s in sheets if s.name == sheet_name]
                if not sheets_to_convert:
                    raise ValueError(f"工作表 '{sheet_name}' 在文件中未找到。")

            # When converting a single sheet from a multi-sheet workbook,
            # the output file name should reflect the sheet name.
            if len(sheets_to_convert) == 1 and len(sheets) > 1:
                 output_p = Path(output_path)
                 final_output_path = str(output_p.parent / f"{output_p.stem}-{sheets_to_convert[0].name}{output_p.suffix or '.html'}")
            else:
                 final_output_path = output_path
                 
            # 检查是否需要分页处理 (分页仅对第一个符合条件的工作表生效)
            if page_size is not None and page_size > 0:
                # 使用分页HTML转换器
                from .converters.paginated_html_converter import PaginatedHTMLConverter
                html_converter = PaginatedHTMLConverter(
                    compact_mode=False,
                    page_size=page_size,
                    page_number=page_number or 1,
                    header_rows=header_rows
                )
                # Paginated converter still works on a single sheet
                result = html_converter.convert_to_file(sheets_to_convert[0], final_output_path)
                return [result] # Return as a list
            else:
                # 使用标准HTML转换器
                html_converter = HTMLConverter(compact_mode=False, header_rows=header_rows)
                results = html_converter.convert_to_files(sheets_to_convert, final_output_path)

            return results

        except Exception as e:
            logger.error(f"HTML转换失败: {e}")
            raise

    def apply_changes(self, file_path: str, table_model_json: dict[str, Any], create_backup: bool = True) -> dict[str, Any]:
        """
        将TableModel JSON的修改应用回原始文件。

        参数：
            file_path: 目标文件路径
            table_model_json: 包含修改的JSON数据
            create_backup: 是否创建备份文件

        返回值：
            操作结果
        """
        try:
            # 验证文件存在
            path = Path(file_path)
            if not path.exists():
                raise FileNotFoundError(f"文件不存在: {file_path}")

            # 验证文件格式
            file_format = path.suffix.lower()
            supported_formats = ['.csv', '.xlsx', '.xlsm', '.xls']
            if file_format not in supported_formats:
                if file_format == '.xlsb':
                    raise ValueError("XLSB格式暂不支持数据写回，请转换为XLSX格式进行编辑")
                raise ValueError(f"Unsupported file type: {file_format}")

            # 创建备份文件（如果需要）
            backup_path = None
            if create_backup:
                backup_path = path.with_suffix(f"{path.suffix}.backup")
                import shutil
                shutil.copy2(path, backup_path)
                logger.info(f"已创建备份文件: {backup_path}")

            # 验证JSON格式
            required_fields = ["sheet_name", "headers", "rows"]
            for field in required_fields:
                if field not in table_model_json:
                    raise ValueError(f"缺少必需字段: {field}")

            # 实现真正的数据写回功能
            changes_applied = 0

            if file_format == '.csv':
                changes_applied = self._write_back_csv(path, table_model_json)
            elif file_format in ['.xlsx', '.xlsm']:
                changes_applied = self._write_back_xlsx(path, table_model_json)
            elif file_format == '.xls':
                changes_applied = self._write_back_xls(path, table_model_json)

            return {
                "status": "success",
                "message": "数据修改已成功应用",
                "file_path": str(path.absolute()),
                "backup_path": str(backup_path) if backup_path else None,
                "backup_created": create_backup,
                "changes_applied": changes_applied,
                "sheet_name": table_model_json.get("sheet_name"),
                "headers_count": len(table_model_json.get("headers", [])),
                "rows_count": len(table_model_json.get("rows", []))
            }

        except Exception as e:
            logger.error(f"应用修改失败: {e}")
            raise

    def _write_back_csv(self, file_path: Path, table_model_json: dict[str, Any]) -> int:
        """
        将修改写回CSV文件。

        参数：
            file_path: CSV文件路径
            table_model_json: 包含修改的JSON数据

        返回值：
            应用的修改数量
        """
        import csv

        headers = table_model_json["headers"]
        rows = table_model_json["rows"]

        # 准备写入的数据
        csv_rows = []

        # 添加表头
        csv_rows.append(headers)

        # 添加数据行
        for row in rows:
            csv_row = []
            for cell in row:
                # 提取单元格值
                if isinstance(cell, dict) and 'value' in cell:
                    value = cell['value']
                else:
                    value = str(cell) if cell is not None else ""
                csv_row.append(str(value) if value is not None else "")
            csv_rows.append(csv_row)

        # 写入CSV文件
        with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerows(csv_rows)

        logger.info(f"CSV文件已更新: {file_path}")
        return len(rows)  # 返回修改的行数

    def _write_back_xls(self, file_path: Path, table_model_json: dict[str, Any]) -> int:
        """
        将修改写回XLS文件。

        参数：
            file_path: XLS文件路径
            table_model_json: 包含修改的JSON数据

        返回值：
            应用的修改数量
        """
        import xlwt

        workbook = xlwt.Workbook(encoding='utf-8')
        sheet_name = table_model_json["sheet_name"]
        worksheet = workbook.add_sheet(sheet_name)

        headers = table_model_json["headers"]
        rows = table_model_json["rows"]

        # 写入表头
        for col_idx, header in enumerate(headers):
            worksheet.write(0, col_idx, header)

        # 写入数据行
        changes_count = 0
        for row_idx, row_data in enumerate(rows, 1):
            for col_idx, cell_data in enumerate(row_data):
                value = cell_data.get('value') if isinstance(cell_data, dict) else cell_data
                worksheet.write(row_idx, col_idx, value)
                changes_count += 1

        workbook.save(str(file_path))
        logger.info(f"XLS文件已更新: {file_path}")
        return changes_count

    def _write_back_xlsx(self, file_path: Path, table_model_json: dict[str, Any]) -> int:
        """
        将修改写回XLSX文件。

        参数：
            file_path: XLSX文件路径
            table_model_json: 包含修改的JSON数据

        返回值：
            应用的修改数量
        """
        import openpyxl

        headers = table_model_json["headers"]
        rows = table_model_json["rows"]

        # 打开现有的工作簿
        workbook = openpyxl.load_workbook(file_path)

        # 根据提供的sheet_name选择正确的工作表
        sheet_name_to_write = table_model_json["sheet_name"]
        
        if sheet_name_to_write in workbook.sheetnames:
            worksheet = workbook[sheet_name_to_write]
        else:
            raise ValueError(f"工作表 '{sheet_name_to_write}' 在文件中不存在。")

        # 清除现有数据（保留样式）
        max_row = worksheet.max_row
        max_col = worksheet.max_column

        # 清除单元格内容但保留格式
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = worksheet.cell(row=row, column=col)
                # 检查是否为合并单元格，跳过MergedCell类型
                try:
                    from openpyxl.cell.cell import MergedCell
                    if isinstance(cell, MergedCell):
                        # 跳过被合并的单元格，因为它们的value属性是只读的
                        continue
                except ImportError:
                    # 如果导入失败，使用字符串检查作为备选方案
                    if 'MergedCell' in str(type(cell)):
                        continue

                cell.value = None

        # 写入表头
        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_idx)

            # 检查是否为合并单元格
            try:
                from openpyxl.cell.cell import MergedCell
                if isinstance(cell, MergedCell):
                    # 跳过被合并的单元格，因为它们的value属性是只读的
                    # 只有合并区域的左上角单元格可以写入
                    logger.debug(f"跳过合并单元格 {cell.coordinate}")
                    continue
            except ImportError:
                # 如果导入失败，使用字符串检查作为备选方案
                if 'MergedCell' in str(type(cell)):
                    logger.debug(f"跳过合并单元格 {cell.coordinate}")
                    continue

            cell.value = header

        # 写入数据行
        changes_count = 0
        for row_idx, row_data in enumerate(rows, 2):  # 从第2行开始（第1行是表头）
            for col_idx, cell_data in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)

                # 提取单元格值
                if isinstance(cell_data, dict) and 'value' in cell_data:
                    value = cell_data['value']
                else:
                    value = cell_data

                # 尝试转换数值类型
                # 检查是否为合并单元格（MergedCell的value属性是只读的）
                is_merged_cell = False
                try:
                    from openpyxl.cell.cell import MergedCell
                    if isinstance(cell, MergedCell):
                        is_merged_cell = True
                except ImportError:
                    # 如果导入失败，使用字符串检查作为备选方案
                    if 'MergedCell' in str(type(cell)):
                        is_merged_cell = True

                if is_merged_cell:
                    # 跳过被合并的单元格，因为它们的value属性是只读的
                    # 只有合并区域的左上角单元格可以写入
                    logger.debug(f"跳过合并单元格 {cell.coordinate}")
                    continue

                # 写入值到普通单元格（此时已确认不是MergedCell）
                try:
                    # 使用类型守卫确保不是 MergedCell
                    from openpyxl.cell.cell import MergedCell
                    if not isinstance(cell, MergedCell):
                        if value is not None and value != "":
                            # 改进的数值转换逻辑
                            if isinstance(value, str):
                                # 尝试转换为数字，使用更安全的方法
                                try:
                                    # 先尝试转换为整数
                                    if '.' not in value and 'e' not in value.lower():
                                        cell.value = int(value)
                                    else:
                                        # 尝试转换为浮点数
                                        cell.value = float(value)
                                except ValueError:
                                    # 如果转换失败，保持为字符串
                                    cell.value = str(value)
                            elif isinstance(value, (int, float, bool)):
                                # 保持原始数据类型
                                cell.value = value
                            else:
                                # 对于其他类型，转换为字符串
                                cell.value = str(value)
                        else:
                            cell.value = None
                except AttributeError as e:
                    # 如果仍然遇到MergedCell问题，记录并跳过
                    if "read-only" in str(e):
                        logger.warning(f"跳过只读单元格 {cell.coordinate}: {e}")
                        continue
                    else:
                        raise

                changes_count += 1

        # 保存工作簿
        workbook.save(file_path)
        workbook.close()

        logger.info(f"XLSX文件已更新: {file_path}")
        return changes_count

    def _sheet_to_json(self, sheet: Sheet, range_string: str | None = None) -> dict[str, Any]:
        """
        将Sheet对象转换为标准化的JSON格式。

        参数：
            sheet: Sheet对象
            range_string: 可选的范围字符串（如"A1:D10"）

        返回值：
            标准化的JSON数据
        """
        # 计算数据大小
        total_cells = self._calculate_data_size(sheet)

        # 智能大小检测 - 针对LLM上下文优化
        current_config = get_config()
        SMALL_FILE_THRESHOLD = current_config.small_file_threshold_cells
        MEDIUM_FILE_THRESHOLD = current_config.medium_file_threshold_cells
        LARGE_FILE_THRESHOLD = current_config.large_file_threshold_cells

        # 处理范围选择
        if range_string:
            try:
                start_row, start_col, end_row, end_col = parse_range_string(range_string)
                return self._extract_range_data(sheet, start_row, start_col, end_row, end_col)
            except ValueError as e:
                logger.warning(f"范围解析失败: {e}, 返回采样数据")
                # 范围解析失败时，返回采样数据而不是完整数据
                return self._extract_sample_data(sheet, total_cells)

        # 根据文件大小选择处理策略 - 更激进的优化
        if total_cells >= LARGE_FILE_THRESHOLD:
            # 大文件：返回摘要
            return self._generate_summary(sheet)
        elif total_cells >= MEDIUM_FILE_THRESHOLD:
            # 中文件：返回采样数据
            return self._extract_sample_data(sheet, total_cells)
        elif total_cells >= SMALL_FILE_THRESHOLD:
            # 小-中文件：返回简化的完整数据（无样式）
            return self._extract_simplified_data(sheet, total_cells)
        else:
            # 小文件：返回完整数据
            full_data = self._extract_full_data(sheet)
            full_data["size_info"] = {
                "total_cells": total_cells,
                "processing_mode": "full",
                "recommendation": "文件较小，已返回完整数据包含样式信息"
            }
            return full_data

    def _extract_optimized_data(self, sheet: Sheet, include_full_data: bool = False,
                               include_styles: bool = False, preview_rows: int = 5,
                               max_rows: int | None = None) -> dict[str, Any]:
        """
        提取优化后的数据，避免上下文爆炸。
        """
        # 基础元数据
        total_rows = len(sheet.rows)
        total_cols = len(sheet.rows[0].cells) if sheet.rows else 0
        total_cells = total_rows * total_cols

        # 提取表头
        headers = []
        if sheet.rows:
            headers = [cell.value if cell is not None and cell.value is not None else f"Column_{i}"
                      for i, cell in enumerate(sheet.rows[0].cells)]

        # 分析数据类型
        data_types = self._analyze_data_types(sheet, headers)

        # 基础响应结构
        response = {
            "sheet_name": sheet.name,
            "metadata": {
                "total_rows": total_rows,
                "total_cols": total_cols,
                "total_cells": total_cells,
                "data_rows": max(0, total_rows - 1),  # 减去表头行
                "has_styles": any(any(cell.style for cell in row.cells if cell is not None) for row in sheet.rows),
                "has_merged_cells": len(sheet.merged_cells) > 0,
                "merged_cells_count": len(sheet.merged_cells),
                "preview_rows": min(preview_rows, max(0, total_rows - 1))
            },
            "headers": headers,
            "data_types": data_types,
            "merged_cells": sheet.merged_cells if len(sheet.merged_cells) <= 10 else sheet.merged_cells[:10]
        }

        # 根据参数决定返回的数据量
        if include_full_data:
            # 返回完整数据
            data_rows = []
            start_row = 1 if total_rows > 1 else 0
            end_row = total_rows

            if max_rows and (end_row - start_row) > max_rows:
                end_row = start_row + max_rows
                response["metadata"]["truncated"] = True
                response["metadata"]["truncated_at"] = max_rows

            for row in sheet.rows[start_row:end_row]:
                row_data = []
                for cell in row.cells:
                    if cell is not None:
                        cell_data = {"value": self._value_to_json_serializable(cell.value)}
                        if include_styles and cell.style:
                            cell_data["style"] = style_to_dict(cell.style)
                    else:
                        cell_data = {"value": None}
                        if include_styles:
                            cell_data["style"] = None
                    row_data.append(cell_data)
                data_rows.append(row_data)

            response["rows"] = data_rows
            response["metadata"]["returned_rows"] = len(data_rows)
        else:
            # 只返回预览数据
            preview_data = []
            start_row = 1 if total_rows > 1 else 0
            end_row = min(start_row + preview_rows, total_rows)

            for row in sheet.rows[start_row:end_row]:
                row_data = []
                for cell in row.cells:
                    # 预览模式下不包含样式，减少数据量
                    if cell is not None:
                        row_data.append(self._value_to_json_serializable(cell.value))
                    else:
                        row_data.append(None)
                preview_data.append(row_data)

            response["preview_rows"] = preview_data

        return response

    def _analyze_data_types(self, sheet: Sheet, headers: list) -> dict[str, str]:
        """分析每列的数据类型。"""
        data_types = {}

        if len(sheet.rows) <= 1:
            return {header: "unknown" for header in headers}

        for col_idx, header in enumerate(headers):
            types_found = set()
            sample_count = 0

            # 检查前几行来推断数据类型
            for row in sheet.rows[1:min(6, len(sheet.rows))]:
                if col_idx < len(row.cells):
                    cell = row.cells[col_idx]
                    # 检查单元格是否为None或者值为None
                    if cell is not None and cell.value is not None:
                        value = cell.value
                        if isinstance(value, str) and value.strip():
                            types_found.add("text")
                        elif isinstance(value, (int, float)):
                            types_found.add("number")
                        elif isinstance(value, bool):
                            types_found.add("boolean")
                        else:
                            types_found.add("other")
                        sample_count += 1

            if len(types_found) == 0:
                data_types[header] = "empty"
            elif len(types_found) == 1:
                data_types[header] = list(types_found)[0]
            else:
                data_types[header] = "mixed"

        return data_types

    def _extract_sample_data(self, sheet: Sheet, total_cells: int) -> dict[str, Any]:
        """提取采样数据，用于中等大小的文件。"""
        # 提取前10行作为样本
        sample_size = min(10, len(sheet.rows))
        sample_rows = []

        # 提取表头
        headers = []
        if sheet.rows:
            headers = [cell.value if cell is not None and cell.value is not None else f"Column_{i}"
                      for i, cell in enumerate(sheet.rows[0].cells)]

        # 提取样本数据
        for i in range(sample_size):
            if i < len(sheet.rows):
                row_data = []
                for cell in sheet.rows[i].cells:
                    if cell is not None:
                        cell_data = {
                            "value": self._value_to_json_serializable(cell.value),
                            "style": style_to_dict(cell.style) if cell.style else None
                        }
                    else:
                        cell_data = {
                            "value": None,
                            "style": None
                        }
                    row_data.append(cell_data)
                sample_rows.append(row_data)

        return {
            "sheet_name": sheet.name,
            "metadata": {
                "total_rows": len(sheet.rows),
                "total_cols": len(sheet.rows[0].cells) if sheet.rows else 0,
                "total_cells": total_cells,
                "processing_mode": "sample",
                "supports_streaming": False
            },
            "sample_data": {
                "headers": headers,
                "rows": sample_rows,
                "note": f"显示前{sample_size}行数据作为样本"
            },
            "size_info": {
                "total_cells": total_cells,
                "processing_mode": "sample",
                "recommendation": f"文件较大({total_cells}个单元格)，已返回采样数据。如需完整数据，请使用范围选择。"
            }
        }

    def _extract_simplified_data(self, sheet: Sheet, total_cells: int) -> dict[str, Any]:
        """提取简化数据，不包含样式信息。"""
        # 提取表头
        headers = []
        data_rows = []

        if sheet.rows:
            # 第一行作为表头
            headers = [cell.value if cell is not None and cell.value is not None else f"Column_{i}"
                      for i, cell in enumerate(sheet.rows[0].cells)]

            # 其余行作为数据（不包含样式）
            for row in sheet.rows[1:]:
                row_data = []
                for cell in row.cells:
                    if cell is not None:
                        row_data.append(self._value_to_json_serializable(cell.value))
                    else:
                        row_data.append(None)
                data_rows.append(row_data)

        return {
            "sheet_name": sheet.name,
            "metadata": {
                "total_rows": len(sheet.rows),
                "total_cols": len(sheet.rows[0].cells) if sheet.rows else 0,
                "total_cells": total_cells,
                "processing_mode": "simplified",
                "supports_streaming": False
            },
            "data": {
                "headers": headers,
                "rows": data_rows
            },
            "size_info": {
                "total_cells": total_cells,
                "processing_mode": "simplified",
                "recommendation": "已返回完整数据但不包含样式信息以节省空间。"
            }
        }

    def _value_to_json_serializable(self, value):
        """Converts a cell value to a JSON serializable format."""
        if isinstance(value, (datetime, date)):
            return value.isoformat()
        if isinstance(value, list):  # Rich text
            return "".join(str(fragment.text) if hasattr(fragment, 'text') else str(fragment) for fragment in value)
        return value

    def _calculate_data_size(self, sheet: Sheet) -> int:
        """计算表格的总单元格数，处理空表格的边界条件。"""
        if not sheet or not sheet.rows:
            return 0
        return sum(len(row.cells) if row and row.cells else 0 for row in sheet.rows)


    def _extract_range_data(self, sheet: Sheet, start_row: int, start_col: int,
                           end_row: int, end_col: int, include_styles: bool = False) -> dict[str, Any]:
        """提取指定范围的数据。"""
        # 验证范围有效性
        if start_row < 0 or start_col < 0:
            raise ValueError("范围起始位置不能为负数")
        if start_row > end_row or start_col > end_col:
            raise ValueError(f"范围无效: 起始位置({start_row},{start_col})不能大于结束位置({end_row},{end_col})")

        # 调整范围以适应实际数据大小
        actual_end_row = min(end_row, len(sheet.rows) - 1)
        if start_row >= len(sheet.rows):
            # 起始行超出数据范围，返回空结果
            return {
                "sheet_name": sheet.name,
                "range": f"{chr(65 + start_col)}{start_row + 1}:{chr(65 + end_col)}{end_row + 1}",
                "headers": [],
                "rows": [],
                "total_rows": 0,
                "total_columns": 0,
                "size_info": {
                    "total_cells": 0,
                    "processing_mode": "range_out_of_bounds",
                    "recommendation": "指定范围超出数据范围"
                }
            }

        # 提取范围内的数据
        range_rows = []
        headers = []

        for row_idx in range(start_row, actual_end_row + 1):
            row = sheet.rows[row_idx]
            row_data = []

            for col_idx in range(start_col, min(end_col + 1, len(row.cells))):
                if col_idx < len(row.cells):
                    cell = row.cells[col_idx]
                    # 检查单元格是否为None
                    if cell is not None:
                        cell_data = {
                            "value": self._value_to_json_serializable(cell.value)
                        }
                        if include_styles:
                            cell_data["style"] = style_to_dict(cell.style) if cell.style else None
                    else:
                        cell_data = {"value": None}
                        if include_styles:
                            cell_data["style"] = None
                else:
                    cell_data = {"value": None}
                    if include_styles:
                        cell_data["style"] = None
                row_data.append(cell_data)

            if row_idx == start_row:
                # 第一行作为表头
                headers = [cell["value"] if cell["value"] is not None else f"Column_{i}"
                          for i, cell in enumerate(row_data)]
            else:
                range_rows.append(row_data)

        return {
            "sheet_name": sheet.name,
            "range": f"{chr(65 + start_col)}{start_row + 1}:{chr(65 + end_col)}{end_row + 1}",
            "metadata": {
                "parser_type": "新解析器系统",
                "range_rows": len(range_rows),
                "range_cols": end_col - start_col + 1,
                "has_styles": any(any(cell.get("style") for cell in row) for row in range_rows)
            },
            "headers": headers,
            "rows": range_rows,
            "size_info": {
                "total_cells": len(range_rows) * (end_col - start_col + 1),
                "processing_mode": "range_selection",
                "recommendation": "已返回指定范围的数据"
            }
        }

    def _generate_summary(self, sheet: Sheet) -> dict[str, Any]:
        """为大文件生成摘要信息。"""
        total_cells = self._calculate_data_size(sheet)

        # 提取前5行作为样本
        sample_rows = []
        headers = []

        for row_idx, row in enumerate(sheet.rows[:5]):
            row_data = []
            for cell in row.cells:
                if cell is not None:
                    cell_data = {
                        "value": self._value_to_json_serializable(cell.value),
                        "style": style_to_dict(cell.style) if cell.style else None
                    }
                else:
                    cell_data = {
                        "value": None,
                        "style": None
                    }
                row_data.append(cell_data)

            if row_idx == 0:
                headers = [cell["value"] if cell["value"] is not None else f"Column_{i}"
                          for i, cell in enumerate(row_data)]
            else:
                sample_rows.append(row_data)

        # 分析数据类型
        data_types = {}
        for col_idx, header in enumerate(headers):
            types_found = set()
            for row in sheet.rows[1:6]:  # 检查前5行数据
                if col_idx < len(row.cells) and row.cells[col_idx].value is not None:
                    value = self._value_to_json_serializable(row.cells[col_idx].value)
                    if isinstance(value, str):
                        types_found.add("text")
                    elif isinstance(value, (int, float)):
                        types_found.add("number")
                    else:
                        types_found.add("other")
            data_types[header] = list(types_found) if types_found else ["unknown"]

        return {
            "sheet_name": sheet.name,
            "metadata": {
                "parser_type": "新解析器系统",
                "total_rows": len(sheet.rows),
                "total_cols": len(headers),
                "total_cells": total_cells,
                "has_styles": any(any(cell.style for cell in row.cells if cell is not None) for row in sheet.rows),
                "data_types": data_types
            },
            "sample_data": {
                "headers": headers,
                "rows": sample_rows,
                "note": f"显示前{len(sample_rows)}行数据作为样本"
            },
            "size_info": {
                "total_cells": total_cells,
                "processing_mode": "summary",
                "recommendation": f"文件过大({total_cells}个单元格)，建议使用范围选择。例如：A1:J100"
            },
            "suggested_ranges": [
                "A1:J100",  # 前100行，前10列
                "A1:Z50",   # 前50行，前26列
                f"A1:{chr(65 + min(25, len(headers) - 1))}200"  # 前200行，实际列数
            ]
        }

    def _extract_full_data(self, sheet: Sheet) -> dict[str, Any]:
        """提取完整的表格数据。"""
        # 提取表头（假设第一行是表头）
        headers = []
        data_rows = []
        
        if sheet.rows:
            # If there's only one row, treat it as data, not a header.
            if len(sheet.rows) == 1:
                 headers = [f"Column_{i}" for i in range(len(sheet.rows[0].cells))]
                 start_row_index = 0
            else:
                 headers = [cell.value if cell is not None and cell.value is not None else f"Column_{i}"
                           for i, cell in enumerate(sheet.rows[0].cells)]
                 start_row_index = 1

            # 提取数据行
            for row in sheet.rows[start_row_index:]:
                row_data = []
                for cell in row.cells:
                    if cell is not None:
                        cell_data = {
                            "value": self._value_to_json_serializable(cell.value),
                            "style": style_to_dict(cell.style) if cell.style else None
                        }
                    else:
                        cell_data = {
                            "value": None,
                            "style": None
                        }
                    row_data.append(cell_data)
                data_rows.append(row_data)

        return {
            "sheet_name": sheet.name,
            "metadata": {
                "parser_type": "新解析器系统",
                "total_rows": len(sheet.rows),
                "total_cols": len(headers),
                "data_rows": len(data_rows),
                "has_styles": any(any(cell.style for cell in row.cells if cell is not None) for row in sheet.rows)
            },
            "headers": headers,
            "rows": data_rows,
            "merged_cells": sheet.merged_cells,
            "format_info": {
                "supports_styles": True,
                "supports_merged_cells": True,
                "supports_hyperlinks": True
            }
        }

    def _should_use_streaming(self, file_path: str, threshold: int) -> bool:
        """
        判断是否应该使用流式读取。
        
        参数:
            file_path: 文件路径
            threshold: 单元格数量阈值
            
        返回:
            如果应该使用流式读取则返回True
        """
        try:
            # 检查解析器是否支持流式读取
            if not ParserFactory.supports_streaming(file_path):
                return False
            
            # 快速估算文件大小
            path = Path(file_path)
            file_size = path.stat().st_size
            
            # 基于文件大小的简单启发式判断
            # 大于配置阈值的文件通常值得流式读取
            current_config = get_config()
            if file_size > current_config.streaming_file_size_mb * 1024 * 1024:
                return True
            
            # 对于较小的文件，可以先快速解析一小部分来估算总大小
            try:
                with StreamingTableReader(file_path) as reader:
                    info = reader.get_info()
                    total_cells = info['total_rows'] * info['total_columns']
                    return total_cells >= threshold
            except Exception as e:
                logger.warning(f"无法估算文件大小: {e}")
                return False
                
        except Exception as e:
            logger.warning(f"检查流式读取支持时出错: {e}")
            return False
    
    def _parse_sheet_streaming(self, file_path: str, sheet_name: str | None = None, 
                              range_string: str | None = None) -> dict[str, Any]:
        """
        使用流式读取器解析表格文件。
        
        参数:
            file_path: 文件路径
            sheet_name: 工作表名称（可选）
            range_string: 单元格范围（可选）

        返回:
            标准化的TableModel JSON
        """
        try:
            with StreamingTableReader(file_path) as reader:
                # 获取文件信息
                file_info = reader.get_info()
                
                # 设置过滤配置
                filter_config = None
                if range_string:
                    filter_config = ChunkFilter(range_string=range_string)
                
                # 检查是否为大文件，如果是则返回摘要
                total_cells = file_info['total_rows'] * file_info['total_columns']
                current_config = get_config()
                LARGE_FILE_THRESHOLD = current_config.streaming_summary_threshold_cells  # 大文件阈值
                
                if total_cells > LARGE_FILE_THRESHOLD and not range_string:
                    # 返回摘要信息
                    return self._generate_streaming_summary(reader, file_info)
                
                # 读取数据块
                all_rows = []
                headers = []
                
                current_config = get_config()
                for chunk in reader.iter_chunks(rows=current_config.streaming_chunk_size_rows, filter_config=filter_config):
                    if not headers:
                        headers = chunk.headers
                    
                    # 转换行格式
                    for row in chunk.rows:
                        row_data = []
                        for cell in row.cells:
                            if cell is not None:
                                cell_data = {
                                    "value": self._value_to_json_serializable(cell.value),
                                    "style": style_to_dict(cell.style) if cell.style else None
                                }
                            else:
                                cell_data = {
                                    "value": None,
                                    "style": None
                                }
                            row_data.append(cell_data)
                        all_rows.append(row_data)
                
                # 构建返回数据
                return {
                    "sheet_name": Path(file_path).stem,
                    "metadata": {
                        "parser_type": file_info['parser_type'],
                        "total_rows": file_info['total_rows'],
                        "total_cols": file_info['total_columns'],
                        "data_rows": len(all_rows),
                        "processing_mode": "streaming",
                        "supports_streaming": True
                    },
                    "headers": headers,
                    "rows": all_rows,
                    "size_info": {
                        "total_cells": total_cells,
                        "processing_mode": "streaming",
                        "recommendation": "使用流式读取处理大文件",
                        "estimated_memory_usage": file_info['estimated_memory_usage']
                    }
                }
                
        except Exception as e:
            logger.error(f"流式解析失败: {e}")
            # 回退到传统方法
            parser = self.parser_factory.get_parser(file_path)
            sheets = parser.parse(file_path)
            # 选择指定的工作表或第一个工作表
            if sheet_name:
                target_sheet = next((s for s in sheets if s.name == sheet_name), None)
                if not target_sheet:
                    raise ValueError(f"工作表 '{sheet_name}' 不存在。")
            else:
                if not sheets:
                    raise ValueError("文件中没有找到任何工作表。")
                target_sheet = sheets[0]
            return self._sheet_to_json(target_sheet, range_string)
    
    def _generate_streaming_summary(self, reader: StreamingTableReader, file_info: dict[str, Any]) -> dict[str, Any]:
        """
        为大文件生成流式摘要信息。
        
        参数:
            reader: 流式读取器
            file_info: 文件信息

        返回:
            摘要数据
        """
        # 读取前几行作为样本
        sample_rows = []
        headers = []
        
        filter_config = ChunkFilter(start_row=0, max_rows=5)
        
        for chunk in reader.iter_chunks(rows=5, filter_config=filter_config):
            headers = chunk.headers
            
            for row in chunk.rows[1:]:  # 跳过表头行
                row_data = []
                for cell in row.cells:
                    if cell is not None:
                        cell_data = {
                            "value": self._value_to_json_serializable(cell.value),
                            "style": style_to_dict(cell.style) if cell.style else None
                        }
                    else:
                        cell_data = {
                            "value": None,
                            "style": None
                        }
                    row_data.append(cell_data)
                sample_rows.append(row_data)
            break  # 只需要第一个块
        
        total_cells = file_info['total_rows'] * file_info['total_columns']
        
        return {
            "sheet_name": Path(reader.file_path).stem,
            "metadata": {
                "parser_type": file_info['parser_type'],
                "total_rows": file_info['total_rows'],
                "total_cols": file_info['total_columns'],
                "total_cells": total_cells,
                "processing_mode": "streaming_summary",
                "supports_streaming": True
            },
            "sample_data": {
                "headers": headers,
                "rows": sample_rows,
                "note": f"显示前{len(sample_rows)}行数据作为样本（流式读取）"
            },
            "size_info": {
                "total_cells": total_cells,
                "processing_mode": "streaming_summary",
                "recommendation": f"文件过大({total_cells}个单元格)，建议使用范围选择。例如：A1:J100",
                "estimated_memory_usage": file_info['estimated_memory_usage']
            },
            "suggested_ranges": [
                "A1:J100",  # 前100行，前10列
                "A1:Z50",   # 前50行，前26列
                f"A1:{chr(65 + min(25, len(headers) - 1))}200"  # 前200行，实际列数
            ]
        }
