"""
高级CSS样式功能的集成测试

测试从Excel文件解析到HTML转换的完整CSS样式处理流程
"""

import pytest
import tempfile
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from pathlib import Path
from src.core_service import CoreService


class TestAdvancedCSSIntegration:
    """高级CSS样式功能的集成测试。"""
    
    def test_xlsx_advanced_styles_end_to_end(self):
        """测试XLSX文件中高级样式的端到端处理。"""
        # 创建包含高级样式的XLSX文件
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        
        # 设置表头
        worksheet['A1'] = 'Product'
        worksheet['B1'] = 'Price'
        worksheet['C1'] = 'Status'
        
        # 创建不同的样式
        # 1. 带边框的标题样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_border = Border(
            top=Side(style="thick"),
            bottom=Side(style="thick"),
            left=Side(style="thick"),
            right=Side(style="thick")
        )
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 应用标题样式
        for cell in ['A1', 'B1', 'C1']:
            worksheet[cell].font = header_font
            worksheet[cell].fill = header_fill
            worksheet[cell].border = header_border
            worksheet[cell].alignment = header_alignment
        
        # 2. 数据行样式
        # 产品名称 - 左对齐，斜体
        worksheet['A2'] = 'Laptop Pro'
        worksheet['A2'].font = Font(italic=True)
        worksheet['A2'].alignment = Alignment(horizontal="left", wrap_text=True)
        
        # 价格 - 右对齐，货币格式
        worksheet['B2'] = 1299.99
        worksheet['B2'].number_format = '$#,##0.00'
        worksheet['B2'].alignment = Alignment(horizontal="right")
        worksheet['B2'].border = Border(left=Side(style="thin"), right=Side(style="thin"))
        
        # 状态 - 居中，背景色
        worksheet['C2'] = 'In Stock'
        worksheet['C2'].fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        worksheet['C2'].alignment = Alignment(horizontal="center")
        
        # 3. 第二行数据 - 不同样式
        worksheet['A3'] = 'Gaming Mouse with very long description that should wrap'
        worksheet['A3'].alignment = Alignment(wrap_text=True)
        worksheet['A3'].border = Border(bottom=Side(style="medium"))
        
        worksheet['B3'] = 89.50
        worksheet['B3'].number_format = '$#,##0.00'
        worksheet['B3'].font = Font(bold=True, color="FF0000")
        
        worksheet['C3'] = 'Out of Stock'
        worksheet['C3'].fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        worksheet['C3'].font = Font(color="8B0000")
        
        # 保存到临时文件
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as xlsx_file:
            xlsx_path = xlsx_file.name
        
        workbook.save(xlsx_path)
        workbook.close()
        
        try:
            service = CoreService()
            
            # 1. 测试解析功能
            print(f"\n=== 测试高级样式解析 ===")
            json_data = service.parse_sheet(xlsx_path)
            
            # 验证解析结果
            assert len(json_data['rows']) == 2  # 数据行，不包括表头
            print(f"解析到的行数: {len(json_data['rows'])}")
            
            # 2. 测试HTML转换功能
            print(f"\n=== 测试高级CSS生成 ===")
            with tempfile.NamedTemporaryFile(suffix='.html', delete=False) as html_file:
                html_path = html_file.name
            
            try:
                html_result = service.convert_to_html(xlsx_path, html_path)
                
                # 验证转换成功
                assert html_result['status'] == 'success'
                assert os.path.exists(html_path)
                
                # 读取HTML内容
                with open(html_path, 'r', encoding='utf-8') as f:
                    html_content = f.read()
                
                print(f"HTML内容长度: {len(html_content)} 字符")
                
                # 验证CSS样式生成
                print(f"\n=== 验证CSS样式 ===")
                
                # 验证边框样式
                assert "border-top:" in html_content or "border:" in html_content
                assert "border-bottom:" in html_content or "border:" in html_content
                print("✅ 边框样式正确生成")
                
                # 验证文本换行
                assert "white-space: pre-wrap" in html_content
                assert "word-wrap: break-word" in html_content
                print("✅ 文本换行样式正确生成")
                
                # 验证数字格式（可能被转换为本地化格式）
                assert "number-format:" in html_content
                assert 'data-number-format=' in html_content
                print("✅ 数字格式正确处理")
                
                # 验证字体样式
                assert "font-weight: bold" in html_content
                assert "font-style: italic" in html_content
                print("✅ 字体样式正确生成")
                
                # 验证颜色样式
                assert "color:" in html_content
                assert "background-color:" in html_content
                print("✅ 颜色样式正确生成")
                
                # 验证对齐样式
                assert "text-align:" in html_content
                assert "vertical-align:" in html_content
                print("✅ 对齐样式正确生成")
                
                # 验证HTML结构
                assert '<table>' in html_content
                assert '<style>' in html_content
                assert 'class=' in html_content  # 样式类应用
                print("✅ HTML结构完整")
                
                # 验证数据内容
                assert 'Laptop Pro' in html_content
                assert '1299.99' in html_content or '1,299.99' in html_content
                assert 'Gaming Mouse' in html_content
                assert 'In Stock' in html_content
                print("✅ 数据内容正确")
                
                print(f"\n✅ 高级CSS样式集成测试成功")
                
            finally:
                if os.path.exists(html_path):
                    os.unlink(html_path)
            
        finally:
            os.unlink(xlsx_path)
    
    def test_css_style_reuse_optimization(self):
        """测试CSS样式复用优化。"""
        # 创建包含重复样式的XLSX文件
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        
        # 创建相同的样式应用到多个单元格
        common_font = Font(bold=True, color="FF0000")
        common_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        common_border = Border(
            top=Side(style="thin"),
            bottom=Side(style="thin"),
            left=Side(style="thin"),
            right=Side(style="thin")
        )
        
        # 应用相同样式到多个单元格
        for row in range(1, 4):
            for col in range(1, 4):
                cell = worksheet.cell(row=row, column=col)
                cell.value = f"Cell {row},{col}"
                cell.font = common_font
                cell.fill = common_fill
                cell.border = common_border
        
        # 保存到临时文件
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as xlsx_file:
            xlsx_path = xlsx_file.name
        
        workbook.save(xlsx_path)
        workbook.close()
        
        try:
            service = CoreService()
            
            # 转换为HTML
            with tempfile.NamedTemporaryFile(suffix='.html', delete=False) as html_file:
                html_path = html_file.name
            
            try:
                html_result = service.convert_to_html(xlsx_path, html_path)
                assert html_result['status'] == 'success'
                
                # 读取HTML内容
                with open(html_path, 'r', encoding='utf-8') as f:
                    html_content = f.read()
                
                # 验证样式复用 - 相同样式应该生成相同的CSS类
                import re
                style_classes = re.findall(r'class="([^"]+)"', html_content)
                
                # 由于所有单元格使用相同样式，应该有很多相同的类名
                if style_classes:
                    most_common_class = max(set(style_classes), key=style_classes.count)
                    class_count = style_classes.count(most_common_class)
                    
                    print(f"最常用的样式类: {most_common_class}, 使用次数: {class_count}")
                    
                    # 应该有多个单元格使用相同的样式类
                    assert class_count >= 3, "样式复用机制应该工作正常"
                
                print(f"✅ CSS样式复用优化测试成功")
                
            finally:
                if os.path.exists(html_path):
                    os.unlink(html_path)
            
        finally:
            os.unlink(xlsx_path)
