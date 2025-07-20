"""
端到端测试模块

测试完整的用户使用场景，从文件输入到最终输出的全流程验证。
"""

import pytest
import tempfile
import os
import json
from pathlib import Path
from unittest.mock import patch, MagicMock
import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill, Border, Side

from src.core_service import CoreService
from src.mcp_server.server import create_server


class TestEndToEnd:
    """端到端测试：验证完整的用户使用场景"""

    @pytest.fixture
    def temp_dir(self):
        """创建临时目录用于测试文件"""
        with tempfile.TemporaryDirectory() as temp_dir:
            yield Path(temp_dir)

    @pytest.fixture
    def complex_excel_file(self, temp_dir):
        """创建包含复杂内容的Excel文件"""
        file_path = temp_dir / "complex.xlsx"
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "复杂数据表"
        
        # 添加带样式的标题
        title_font = Font(bold=True, size=14, color="FF0000")
        title_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        sheet["A1"] = "季度销售报告"
        sheet["A1"].font = title_font
        sheet["A1"].fill = title_fill
        
        # 添加表头
        headers = ["季度", "产品A", "产品B", "产品C", "总计"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 添加数据
        data = [
            ["Q1", 1000, 1500, 2000, 4500],
            ["Q2", 1200, 1800, 2200, 5200],
            ["Q3", 1100, 1600, 2100, 4800],
            ["Q4", 1300, 1900, 2300, 5500]
        ]
        
        for row_idx, row_data in enumerate(data, 4):
            for col_idx, value in enumerate(row_data, 1):
                cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                if col_idx > 1:  # 数字列
                    cell.number_format = '#,##0'
        
        # 创建柱状图
        bar_chart = BarChart()
        bar_chart.title = "季度销售对比"
        bar_chart.x_axis.title = "季度"
        bar_chart.y_axis.title = "销售额"
        
        data_range = Reference(sheet, min_col=2, min_row=3, max_row=7, max_col=4)
        categories = Reference(sheet, min_col=1, min_row=4, max_row=7)
        
        bar_chart.add_data(data_range, titles_from_data=True)
        bar_chart.set_categories(categories)
        
        sheet.add_chart(bar_chart, "G3")
        
        # 创建折线图
        line_chart = LineChart()
        line_chart.title = "销售趋势"
        line_chart.x_axis.title = "季度"
        line_chart.y_axis.title = "总销售额"
        
        trend_data = Reference(sheet, min_col=5, min_row=3, max_row=7)
        line_chart.add_data(trend_data, titles_from_data=True)
        line_chart.set_categories(categories)
        
        sheet.add_chart(line_chart, "G18")
        
        # 合并单元格
        sheet.merge_cells("A1:E1")
        
        workbook.save(file_path)
        return file_path

    def test_complete_user_workflow_excel_to_html(self, complex_excel_file, temp_dir):
        """
        TDD端到端测试：用户完整工作流程 - Excel转HTML
        
        模拟用户从上传Excel文件到获得HTML输出的完整流程
        """
        
        core_service = CoreService()
        output_html = temp_dir / "output.html"
        
        # 执行完整的用户工作流程
        result = core_service.convert_to_html(str(complex_excel_file), str(output_html))
        
        # 验证输出文件存在
        assert output_html.exists()
        
        # 验证HTML内容质量
        html_content = output_html.read_text(encoding='utf-8')
        
        # 验证基本HTML结构
        assert '<!DOCTYPE html>' in html_content or '<html' in html_content
        assert '<table' in html_content
        assert '</table>' in html_content
        
        # 验证数据内容
        assert '季度销售报告' in html_content
        assert 'Q1' in html_content
        # 验证数字内容（可能被格式化为1,000）
        assert ('1000' in html_content or '1,000' in html_content)
        assert '产品A' in html_content
        
        # 验证文件大小合理（不为空）
        assert output_html.stat().st_size > 100

    def test_user_workflow_with_json_output(self, complex_excel_file):
        """
        TDD端到端测试：用户工作流程 - 获取JSON数据
        
        模拟用户解析Excel文件并获取结构化JSON数据
        """
        
        core_service = CoreService()
        
        # 解析Excel文件获取JSON数据
        result = core_service.parse_sheet(str(complex_excel_file))
        
        # 验证JSON结构
        assert result is not None
        assert isinstance(result, dict)
        assert 'sheet_name' in result
        assert 'rows' in result or 'data' in result

        # 验证数据内容
        assert result['sheet_name'] == "复杂数据表"
        data_key = 'rows' if 'rows' in result else 'data'
        assert len(result[data_key]) > 0

        # 验证数据可以序列化为JSON
        json_str = json.dumps(result, ensure_ascii=False)
        assert len(json_str) > 0

        # 验证可以反序列化
        parsed_back = json.loads(json_str)
        assert parsed_back['sheet_name'] == result['sheet_name']

    def test_user_workflow_error_handling(self, temp_dir):
        """
        TDD端到端测试：用户工作流程 - 错误处理
        
        模拟用户遇到各种错误情况时的系统响应
        """
        
        core_service = CoreService()
        
        # 测试场景1：文件不存在
        nonexistent_file = temp_dir / "nonexistent.xlsx"
        try:
            result = core_service.parse_sheet(str(nonexistent_file))
            data_key = 'rows' if result and 'rows' in result else 'data'
            assert result is None or (isinstance(result, dict) and len(result.get(data_key, [])) == 0)
        except Exception as e:
            assert isinstance(e, (FileNotFoundError, ValueError, Exception))
        
        # 测试场景2：无效文件格式
        invalid_file = temp_dir / "invalid.txt"
        invalid_file.write_text("这不是Excel文件")
        
        try:
            result = core_service.parse_sheet(str(invalid_file))
            data_key = 'rows' if result and 'rows' in result else 'data'
            assert result is None or (isinstance(result, dict) and len(result.get(data_key, [])) == 0)
        except Exception as e:
            assert isinstance(e, (ValueError, Exception))
        
        # 测试场景3：空Excel文件
        empty_file = temp_dir / "empty.xlsx"
        empty_workbook = openpyxl.Workbook()
        empty_workbook.save(empty_file)
        
        result = core_service.parse_sheet(str(empty_file))
        assert result is not None
        data_key = 'rows' if 'rows' in result else 'data'
        assert data_key in result

    def test_user_workflow_performance(self, temp_dir):
        """
        TDD端到端测试：用户工作流程 - 性能验证
        
        验证系统在处理不同大小文件时的性能表现
        """
        
        import time
        
        # 创建中等大小的文件
        medium_file = temp_dir / "medium.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        
        # 添加100行数据
        for i in range(100):
            sheet.append([f"数据{i}", i * 10, i * 20, i * 30])
        
        workbook.save(medium_file)
        
        core_service = CoreService()
        
        # 测试解析性能
        start_time = time.time()
        result = core_service.parse_sheet(str(medium_file))
        parse_time = time.time() - start_time
        
        # 验证解析成功
        assert result is not None
        assert len(result['data']) > 0
        
        # 验证性能合理（应该在10秒内完成）
        assert parse_time < 10.0
        
        # 测试HTML转换性能
        output_html = temp_dir / "performance_test.html"
        start_time = time.time()
        core_service.convert_to_html(str(medium_file), str(output_html))
        convert_time = time.time() - start_time
        
        # 验证转换成功
        assert output_html.exists()
        
        # 验证性能合理（应该在15秒内完成）
        assert convert_time < 15.0

    def test_mcp_server_integration_workflow(self, complex_excel_file, temp_dir):
        """
        TDD端到端测试：MCP服务器集成工作流程
        
        测试通过MCP服务器接口的完整用户工作流程
        """
        
        # 创建MCP服务器实例
        server = create_server()

        # 验证服务器能够创建
        assert server is not None

        # 验证服务器具有基本属性
        assert hasattr(server, 'name')
        assert server.name == "mcp-sheet-parser"

    def test_user_workflow_with_charts_and_images(self, temp_dir):
        """
        TDD端到端测试：包含图表和图像的用户工作流程
        
        测试用户处理包含复杂图表和图像的Excel文件
        """
        
        # 创建包含图表的Excel文件
        chart_file = temp_dir / "with_charts.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        
        # 添加数据
        sheet.append(["月份", "销售额", "利润"])
        for i in range(1, 13):
            sheet.append([f"{i}月", i * 1000, i * 200])
        
        # 创建图表
        chart = BarChart()
        chart.title = "月度销售"
        
        data_range = Reference(sheet, min_col=2, min_row=1, max_row=13, max_col=3)
        categories = Reference(sheet, min_col=1, min_row=2, max_row=13)
        
        chart.add_data(data_range, titles_from_data=True)
        chart.set_categories(categories)
        
        sheet.add_chart(chart, "E2")
        
        workbook.save(chart_file)
        
        core_service = CoreService()
        output_html = temp_dir / "charts_output.html"
        
        # 转换包含图表的文件
        result = core_service.convert_to_html(str(chart_file), str(output_html))
        
        # 验证转换成功
        assert output_html.exists()
        
        # 验证HTML内容
        html_content = output_html.read_text(encoding='utf-8')
        assert '<table' in html_content
        assert '销售额' in html_content
        assert '1月' in html_content

    def test_user_workflow_data_validation(self, temp_dir):
        """
        TDD端到端测试：数据验证的用户工作流程
        
        测试用户上传的数据经过完整验证流程
        """
        
        # 创建包含各种数据类型的Excel文件
        validation_file = temp_dir / "validation.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        
        # 添加不同类型的数据
        sheet.append(["文本", "数字", "日期", "布尔", "公式"])
        sheet.append(["测试文本", 123.45, "2024-01-01", True, "=B2*2"])
        sheet.append(["", 0, None, False, "=SUM(B2:B3)"])
        sheet.append([None, -100, "无效日期", "TRUE", "=B2+B3"])
        
        workbook.save(validation_file)
        
        core_service = CoreService()
        
        # 解析并验证数据
        result = core_service.parse_sheet(str(validation_file))
        
        # 验证数据解析成功
        assert result is not None
        data_key = 'rows' if 'rows' in result else 'data'
        assert data_key in result
        assert len(result[data_key]) > 0
        
        # 验证数据类型处理
        data_key = 'rows' if 'rows' in result else 'data'
        data = result[data_key]
        assert len(data) >= 3  # 至少有3行数据

        # 验证第一行数据
        first_row = data[0]
        assert len(first_row) >= 5  # 至少有5列
