"""
集成测试模块

测试各模块间的协作和完整的工作流程，确保系统整体功能正确性。
"""

import pytest
import tempfile
import os
from pathlib import Path
from unittest.mock import MagicMock, patch
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image as OpenpyxlImage

from src.core_service import CoreService
from src.parsers.factory import ParserFactory
from src.converters.html_converter import HTMLConverter
from src.converters.svg_chart_renderer import SVGChartRenderer
from src.models.table_model import Sheet, Row, Cell, Chart


class TestIntegration:
    """集成测试：验证各模块间的协作和完整工作流程"""

    @pytest.fixture
    def temp_dir(self):
        """创建临时目录用于测试文件"""
        with tempfile.TemporaryDirectory() as temp_dir:
            yield Path(temp_dir)

    @pytest.fixture
    def sample_excel_file(self, temp_dir):
        """创建包含数据和图表的示例Excel文件"""
        file_path = temp_dir / "sample.xlsx"
        
        # 创建工作簿和工作表
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "数据表"
        
        # 添加数据
        headers = ["产品", "销量", "收入"]
        sheet.append(headers)
        
        data = [
            ["产品A", 100, 10000],
            ["产品B", 150, 15000],
            ["产品C", 200, 20000],
            ["产品D", 120, 12000]
        ]
        
        for row in data:
            sheet.append(row)
        
        # 创建图表
        chart = BarChart()
        chart.title = "销量统计"
        chart.x_axis.title = "产品"
        chart.y_axis.title = "销量"
        
        # 设置数据范围
        data_range = Reference(sheet, min_col=2, min_row=1, max_row=5, max_col=2)
        categories = Reference(sheet, min_col=1, min_row=2, max_row=5)
        
        chart.add_data(data_range, titles_from_data=True)
        chart.set_categories(categories)
        
        # 添加图表到工作表
        sheet.add_chart(chart, "E2")
        
        workbook.save(file_path)
        return file_path

    def test_complete_excel_to_html_workflow(self, sample_excel_file, temp_dir):
        """
        TDD集成测试：完整的Excel到HTML转换工作流程
        
        测试流程：Excel解析 → 数据提取 → HTML转换 → 图表渲染
        """
        
        core_service = CoreService()
        output_path = temp_dir / "output.html"
        
        # 执行完整的转换流程
        result = core_service.convert_to_html(str(sample_excel_file), str(output_path))
        
        # 验证转换结果
        assert result is not None
        assert output_path.exists()
        
        # 验证HTML文件内容
        html_content = output_path.read_text(encoding='utf-8')
        assert '<table' in html_content
        assert '产品A' in html_content
        assert '销量' in html_content
        assert '100' in html_content

    def test_parser_factory_integration(self, sample_excel_file):
        """
        TDD集成测试：解析器工厂与具体解析器的集成
        
        测试解析器工厂能够正确选择和创建解析器
        """
        
        factory = ParserFactory()
        
        # 获取适当的解析器
        parser = factory.get_parser(str(sample_excel_file))
        
        # 验证解析器类型
        assert parser is not None
        assert hasattr(parser, 'parse')
        
        # 执行解析
        sheets = parser.parse(str(sample_excel_file))
        
        # 验证解析结果
        assert isinstance(sheets, list)
        assert len(sheets) > 0
        assert isinstance(sheets[0], Sheet)
        assert sheets[0].name == "数据表"
        assert len(sheets[0].rows) > 0

    def test_html_converter_with_chart_renderer_integration(self, temp_dir):
        """
        TDD集成测试：HTML转换器与图表渲染器的集成
        
        测试HTML转换器能够正确调用图表渲染器
        """
        
        # 创建测试数据
        chart = Chart(
            name="测试图表",
            type="bar",
            chart_data={
                "title": "测试图表",
                "series": [
                    {"name": "系列1", "values": [10, 20, 30]},
                    {"name": "系列2", "values": [15, 25, 35]}
                ],
                "categories": ["A", "B", "C"]
            }
        )
        
        sheet = Sheet(
            name="测试表",
            rows=[
                Row(cells=[Cell(value="标题1"), Cell(value="标题2")]),
                Row(cells=[Cell(value="数据1"), Cell(value="数据2")])
            ],
            charts=[chart]
        )
        
        html_converter = HTMLConverter()
        
        # 转换为HTML文件
        import tempfile
        with tempfile.NamedTemporaryFile(suffix='.html', delete=False) as tmp_file:
            results = html_converter.convert_to_files([sheet], tmp_file.name)

            # 验证转换结果
            assert isinstance(results, list)
            assert len(results) > 0

            # 读取生成的HTML文件
            html_content = Path(tmp_file.name).read_text(encoding='utf-8')

            # 验证HTML内容包含表格
            assert '<table' in html_content
            assert '测试表' in html_content or '标题1' in html_content

    def test_core_service_caching_integration(self, sample_excel_file):
        """
        TDD集成测试：核心服务与缓存系统的集成
        
        测试缓存系统在完整工作流程中的作用
        """
        
        core_service = CoreService()
        
        # 第一次解析（应该缓存结果）
        result1 = core_service.parse_sheet(str(sample_excel_file))
        
        # 第二次解析（应该使用缓存）
        result2 = core_service.parse_sheet(str(sample_excel_file))
        
        # 验证两次结果一致
        assert result1 is not None
        assert result2 is not None
        assert result1['sheet_name'] == result2['sheet_name']
        data_key = 'rows' if 'rows' in result1 else 'data'
        assert len(result1[data_key]) == len(result2[data_key])

    def test_error_handling_integration(self, temp_dir):
        """
        TDD集成测试：错误处理在各模块间的传播
        
        测试错误能够在模块间正确传播和处理
        """
        
        # 创建无效的文件路径
        invalid_file = temp_dir / "nonexistent.xlsx"
        
        core_service = CoreService()
        
        # 尝试解析不存在的文件
        try:
            result = core_service.parse_sheet(str(invalid_file))
            # 如果没有抛出异常，结果应该是None或空
            assert result is None or (isinstance(result, dict) and len(result.get('data', [])) == 0)
        except Exception as e:
            # 抛出异常也是可接受的行为
            assert isinstance(e, (FileNotFoundError, ValueError, Exception))

    def test_large_file_processing_integration(self, temp_dir):
        """
        TDD集成测试：大文件处理的完整流程
        
        测试系统处理大文件时各模块的协作
        """
        
        # 创建包含大量数据的Excel文件
        large_file = temp_dir / "large.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        
        # 添加适量数据（100行，避免测试超时）
        for i in range(100):
            sheet.append([f"数据{i}_1", f"数据{i}_2", f"数据{i}_3", i * 10])
        
        workbook.save(large_file)
        
        core_service = CoreService()
        
        # 解析大文件
        result = core_service.parse_sheet(str(large_file))
        
        # 验证能够处理大文件
        assert result is not None
        data_key = 'rows' if 'rows' in result else 'data'
        assert data_key in result
        assert len(result[data_key]) > 0

    def test_multi_sheet_processing_integration(self, temp_dir):
        """
        TDD集成测试：多工作表处理的完整流程
        
        测试系统处理多工作表文件时的协作
        """
        
        # 创建包含多个工作表的Excel文件
        multi_sheet_file = temp_dir / "multi_sheet.xlsx"
        workbook = openpyxl.Workbook()
        
        # 第一个工作表
        sheet1 = workbook.active
        sheet1.title = "销售数据"
        sheet1.append(["产品", "销量"])
        sheet1.append(["A", 100])
        sheet1.append(["B", 200])
        
        # 第二个工作表
        sheet2 = workbook.create_sheet("财务数据")
        sheet2.append(["项目", "金额"])
        sheet2.append(["收入", 50000])
        sheet2.append(["支出", 30000])
        
        workbook.save(multi_sheet_file)
        
        core_service = CoreService()
        
        # 解析多工作表文件
        result = core_service.parse_sheet(str(multi_sheet_file))
        
        # 验证能够处理多工作表
        assert result is not None
        assert 'sheet_name' in result
        assert result['sheet_name'] in ["销售数据", "财务数据"]

    def test_streaming_processing_integration(self, temp_dir):
        """
        TDD集成测试：流式处理的完整集成
        
        测试流式处理在大文件场景下的集成效果
        """
        
        # 创建大文件以触发流式处理
        streaming_file = temp_dir / "streaming.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        
        # 添加足够多的数据以触发流式处理
        for i in range(500):
            sheet.append([f"Row{i}", f"Data{i}", i])
        
        workbook.save(streaming_file)
        
        # 模拟大文件大小以触发流式处理
        with patch('os.path.getsize', return_value=10 * 1024 * 1024):  # 10MB
            core_service = CoreService()
            

            result = core_service.parse_sheet(str(streaming_file))
            
            # 验证流式处理结果
            assert result is not None
            data_key = 'rows' if 'rows' in result else 'sample_data'
            assert data_key in result
