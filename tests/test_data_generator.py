"""
Test data generator for comprehensive chart testing.

This module generates various Excel chart types and configurations
for testing chart processing components.
"""

import openpyxl
from openpyxl.chart import (
    BarChart, LineChart, PieChart, AreaChart, ScatterChart,
    Reference
)
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.colors import ColorChoice, SchemeColor, RGBPercent
from openpyxl.chart.data_source import NumData, NumVal, StrData, StrVal
from pathlib import Path
import random


class ChartTestDataGenerator:
    """Generate test data for chart processing tests."""
    
    def __init__(self, output_dir="tests/test_files"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
    
    def create_basic_bar_chart_file(self):
        """Create Excel file with basic bar chart."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bar Chart Data"
        
        # Add data
        data = [
            ["Category", "Values"],
            ["A", 10],
            ["B", 25],
            ["C", 15],
            ["D", 30],
            ["E", 20]
        ]
        
        for row in data:
            ws.append(row)
        
        # Create bar chart
        chart = BarChart()
        chart.title = "Basic Bar Chart"
        chart.x_axis.title = "Categories"
        chart.y_axis.title = "Values"
        
        # Add data to chart
        data_ref = Reference(ws, min_col=2, min_row=1, max_row=6, max_col=2)
        cats_ref = Reference(ws, min_col=1, min_row=2, max_row=6, max_col=1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        
        # Position chart
        ws.add_chart(chart, "D2")
        
        file_path = self.output_dir / "basic_bar_chart.xlsx"
        wb.save(file_path)
        return file_path
    
    def create_multi_series_line_chart_file(self):
        """Create Excel file with multi-series line chart."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Line Chart Data"
        
        # Add data
        data = [
            ["Month", "Series 1", "Series 2", "Series 3"],
            ["Jan", 10, 15, 8],
            ["Feb", 15, 20, 12],
            ["Mar", 12, 18, 15],
            ["Apr", 20, 25, 18],
            ["May", 18, 22, 20],
            ["Jun", 25, 30, 25]
        ]
        
        for row in data:
            ws.append(row)
        
        # Create line chart
        chart = LineChart()
        chart.title = "Multi-Series Line Chart"
        chart.x_axis.title = "Months"
        chart.y_axis.title = "Values"
        
        # Add data series
        for col in range(2, 5):  # Columns B, C, D
            data_ref = Reference(ws, min_col=col, min_row=1, max_row=7, max_col=col)
            chart.add_data(data_ref, titles_from_data=True)
        
        cats_ref = Reference(ws, min_col=1, min_row=2, max_row=7, max_col=1)
        chart.set_categories(cats_ref)
        
        # Position chart
        ws.add_chart(chart, "F2")
        
        file_path = self.output_dir / "multi_series_line_chart.xlsx"
        wb.save(file_path)
        return file_path
    
    def create_styled_pie_chart_file(self):
        """Create Excel file with styled pie chart."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pie Chart Data"
        
        # Add data
        data = [
            ["Category", "Values"],
            ["Product A", 30],
            ["Product B", 25],
            ["Product C", 20],
            ["Product D", 15],
            ["Product E", 10]
        ]
        
        for row in data:
            ws.append(row)
        
        # Create pie chart
        chart = PieChart()
        chart.title = "Styled Pie Chart"
        
        # Add data to chart
        data_ref = Reference(ws, min_col=2, min_row=1, max_row=6, max_col=2)
        cats_ref = Reference(ws, min_col=1, min_row=2, max_row=6, max_col=1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        
        # Style the chart
        chart.legend.position = 'r'  # Right position
        
        # Position chart
        ws.add_chart(chart, "D2")
        
        file_path = self.output_dir / "styled_pie_chart.xlsx"
        wb.save(file_path)
        return file_path
    
    def create_area_chart_file(self):
        """Create Excel file with area chart."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Area Chart Data"
        
        # Add data
        data = [
            ["Quarter", "Revenue", "Costs"],
            ["Q1", 100, 60],
            ["Q2", 120, 70],
            ["Q3", 110, 65],
            ["Q4", 140, 80]
        ]
        
        for row in data:
            ws.append(row)
        
        # Create area chart
        chart = AreaChart()
        chart.title = "Revenue vs Costs Area Chart"
        chart.x_axis.title = "Quarters"
        chart.y_axis.title = "Amount ($000)"
        
        # Add data series
        for col in range(2, 4):  # Columns B, C
            data_ref = Reference(ws, min_col=col, min_row=1, max_row=5, max_col=col)
            chart.add_data(data_ref, titles_from_data=True)
        
        cats_ref = Reference(ws, min_col=1, min_row=2, max_row=5, max_col=1)
        chart.set_categories(cats_ref)
        
        # Position chart
        ws.add_chart(chart, "E2")
        
        file_path = self.output_dir / "area_chart.xlsx"
        wb.save(file_path)
        return file_path
    
    def create_complex_chart_file(self):
        """Create Excel file with complex chart features."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Complex Chart Data"
        
        # Add comprehensive data
        data = [
            ["Product", "Q1 Sales", "Q2 Sales", "Q3 Sales", "Q4 Sales", "Target"],
            ["Product A", 85, 92, 88, 95, 90],
            ["Product B", 78, 85, 82, 89, 85],
            ["Product C", 92, 88, 95, 98, 95],
            ["Product D", 65, 72, 68, 75, 70],
            ["Product E", 88, 95, 92, 99, 95]
        ]
        
        for row in data:
            ws.append(row)
        
        # Create bar chart with multiple series
        chart = BarChart()
        chart.title = "Complex Sales Analysis Chart"
        chart.x_axis.title = "Products"
        chart.y_axis.title = "Sales Volume"
        
        # Add quarterly data
        for col in range(2, 6):  # Q1-Q4 columns
            data_ref = Reference(ws, min_col=col, min_row=1, max_row=6, max_col=col)
            chart.add_data(data_ref, titles_from_data=True)
        
        cats_ref = Reference(ws, min_col=1, min_row=2, max_row=6, max_col=1)
        chart.set_categories(cats_ref)
        
        # Style the chart
        chart.legend.position = 'b'  # Bottom position
        chart.height = 10  # Increase height
        chart.width = 15   # Increase width
        
        # Position chart
        ws.add_chart(chart, "H2")
        
        file_path = self.output_dir / "complex_chart.xlsx"
        wb.save(file_path)
        return file_path
    
    def create_large_dataset_chart_file(self):
        """Create Excel file with large dataset chart."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Large Dataset"
        
        # Generate large dataset (100 data points)
        ws.append(["Index", "Value 1", "Value 2"])
        
        for i in range(1, 101):
            value1 = random.randint(50, 150) + random.random() * 10
            value2 = random.randint(30, 120) + random.random() * 15
            ws.append([f"Point_{i}", value1, value2])
        
        # Create line chart
        chart = LineChart()
        chart.title = "Large Dataset Performance Chart"
        chart.x_axis.title = "Data Points"
        chart.y_axis.title = "Values"
        
        # Add data series
        for col in range(2, 4):  # Columns B, C
            data_ref = Reference(ws, min_col=col, min_row=1, max_row=101, max_col=col)
            chart.add_data(data_ref, titles_from_data=True)
        
        cats_ref = Reference(ws, min_col=1, min_row=2, max_row=101, max_col=1)
        chart.set_categories(cats_ref)
        
        # Position chart
        ws.add_chart(chart, "E2")
        
        file_path = self.output_dir / "large_dataset_chart.xlsx"
        wb.save(file_path)
        return file_path
    
    def create_edge_case_chart_file(self):
        """Create Excel file with edge case data."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Edge Cases"
        
        # Add edge case data
        data = [
            ["Category", "Values", "Negative", "Zero"],
            ["Normal", 100, -50, 0],
            ["Large", 10000, -5000, 0],
            ["Small", 0.1, -0.05, 0],
            ["Zero", 0, 0, 0],
            ["Negative", -100, -200, 0]
        ]
        
        for row in data:
            ws.append(row)
        
        # Create bar chart with mixed positive/negative values
        chart = BarChart()
        chart.title = "Edge Case Data Chart"
        chart.x_axis.title = "Categories"
        chart.y_axis.title = "Values"
        
        # Add data series
        for col in range(2, 5):  # Columns B, C, D
            data_ref = Reference(ws, min_col=col, min_row=1, max_row=6, max_col=col)
            chart.add_data(data_ref, titles_from_data=True)
        
        cats_ref = Reference(ws, min_col=1, min_row=2, max_row=6, max_col=1)
        chart.set_categories(cats_ref)
        
        # Position chart
        ws.add_chart(chart, "F2")
        
        file_path = self.output_dir / "edge_case_chart.xlsx"
        wb.save(file_path)
        return file_path
    
    def create_unicode_chart_file(self):
        """Create Excel file with Unicode characters."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Unicode Data"
        
        # Add Unicode data
        data = [
            ["产品", "销售额"],  # Chinese
            ["产品A", 1000],
            ["产品B", 1500],
            ["منتج", 1200],    # Arabic
            ["Продукт", 800],  # Russian
            ["Produit", 1100]  # French
        ]
        
        for row in data:
            ws.append(row)
        
        # Create pie chart
        chart = PieChart()
        chart.title = "国际销售分析"  # International Sales Analysis in Chinese
        
        # Add data to chart
        data_ref = Reference(ws, min_col=2, min_row=1, max_row=6, max_col=2)
        cats_ref = Reference(ws, min_col=1, min_row=2, max_row=6, max_col=1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        
        # Position chart
        ws.add_chart(chart, "D2")
        
        file_path = self.output_dir / "unicode_chart.xlsx"
        wb.save(file_path)
        return file_path
    
    def create_all_test_files(self):
        """Create all test files and return their paths."""
        test_files = []
        
        print("Generating test chart files...")
        
        test_files.append(self.create_basic_bar_chart_file())
        print("✓ Basic bar chart created")
        
        test_files.append(self.create_multi_series_line_chart_file())
        print("✓ Multi-series line chart created")
        
        test_files.append(self.create_styled_pie_chart_file())
        print("✓ Styled pie chart created")
        
        test_files.append(self.create_area_chart_file())
        print("✓ Area chart created")
        
        test_files.append(self.create_complex_chart_file())
        print("✓ Complex chart created")
        
        test_files.append(self.create_large_dataset_chart_file())
        print("✓ Large dataset chart created")
        
        test_files.append(self.create_edge_case_chart_file())
        print("✓ Edge case chart created")
        
        test_files.append(self.create_unicode_chart_file())
        print("✓ Unicode chart created")
        
        print(f"\nGenerated {len(test_files)} test files in {self.output_dir}")
        return test_files


class TestDataValidator:
    """Validate generated test data."""
    
    @staticmethod
    def validate_chart_file(file_path):
        """Validate that a chart file contains expected elements."""
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # Check that file has data
            assert ws.max_row > 1, "File should have data rows"
            assert ws.max_column > 1, "File should have multiple columns"
            
            # Check for charts
            assert len(ws._charts) > 0, "File should contain at least one chart"
            
            chart = ws._charts[0]
            assert chart.title is not None, "Chart should have a title"
            
            return True
            
        except Exception as e:
            print(f"Validation failed for {file_path}: {e}")
            return False
    
    @staticmethod
    def validate_all_test_files(file_paths):
        """Validate all generated test files."""
        valid_files = []
        invalid_files = []
        
        for file_path in file_paths:
            if TestDataValidator.validate_chart_file(file_path):
                valid_files.append(file_path)
            else:
                invalid_files.append(file_path)
        
        print(f"\nValidation Results:")
        print(f"✓ Valid files: {len(valid_files)}")
        print(f"✗ Invalid files: {len(invalid_files)}")
        
        if invalid_files:
            print("Invalid files:")
            for file_path in invalid_files:
                print(f"  - {file_path}")
        
        return valid_files, invalid_files


if __name__ == "__main__":
    # Generate all test files
    generator = ChartTestDataGenerator()
    test_files = generator.create_all_test_files()
    
    # Validate generated files
    validator = TestDataValidator()
    valid_files, invalid_files = validator.validate_all_test_files(test_files)
    
    if len(valid_files) == len(test_files):
        print("\n🎉 All test files generated and validated successfully!")
    else:
        print(f"\n⚠️  {len(invalid_files)} files failed validation")